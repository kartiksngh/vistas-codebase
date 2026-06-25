"""
funds_portfolio.py — mutual-fund PORTFOLIO HOLDINGS (look-through) for Vistas Active Phase 2.

WHAT THIS IS (first principles)
-------------------------------
A fund's NAV tells you the *result*; the PORTFOLIO tells you the *cause* — what the manager
actually owns. Once you can see holdings you can answer the questions that matter: what is the
fund really exposed to (sector / asset mix), how concentrated is it (top-10, Herfindahl), how
much does it overlap a benchmark or another fund, and how fast does it churn. This module is
the look-through layer beneath the NAV layer (funds_nav.py).

SOURCE
------
Per-AMC MONTHLY portfolio-disclosure workbooks (SEBI Reg. 59A; the MFD/CIR/9/120/2000 standard
sheet). One workbook per AMC holds ALL its schemes — one worksheet per scheme — so it is ~40
AMC files a month, not thousands of per-scheme fetches. These are plain static-CDN GETs (no
Akamai/WAF), so this runs head-less / from a datacenter IP. Verified live: Groww + Nippon.

The SEBI sheet (verified against Groww, 2026-05-31):
  row: scheme name ("IB01-Groww Large Cap Fund")  ·  row: "Portfolio as on 31-MAY-2026"
  HEADER: ISIN | Name of Instrument | Rating/Industry | Quantity | Market Value (In Rs. lakh)
          | % To Net Assets | Maturity Date | Put/Call Option
  then SECTION blocks (EQUITY & EQUITY RELATED / DEBT / MONEY MARKET / TREPS / Net
  Receivables…) of holding rows, ending in a "Grand Total" row.
GOTCHA: the "% To Net Assets" column is sometimes a FRACTION (Groww: 0.0822 = 8.22 %, grand
total 1.0) and sometimes percent-points (others: 8.22). We DETECT per sheet (grand-total or the
holdings sum) and normalise everything to PERCENT. Columns can be shifted (Groww leaves col A
blank), so we locate every field BY HEADER TEXT, never by a fixed index.

DISPLAY-PLANE: like macro.py / stock_intel.py, every aggregate (asset mix, sector mix, top
holdings, concentration) is precomputed HERE in Python and baked into one small per-scheme
JSON; the browser only renders. No JS-parity port — analytics.py is untouched.

Graceful-degrade contract: one AMC failing never blocks the others; a total failure keeps the
cached snapshot. Network is OPTIONAL.
"""
from __future__ import annotations

import os
import io
import re
import html
import json
import time
import datetime as dt

try:
    import requests
except Exception:                                   # pragma: no cover
    requests = None

try:
    from openpyxl import load_workbook
except Exception:                                   # pragma: no cover
    load_workbook = None

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_OUT_DEFAULT = os.path.join(_ROOT, "data", "funds", "portfolio")
# Raw-workbook byte cache. The AMC files are MONTHLY, so once a month's workbook is downloaded
# every later rebuild reads it from disk instead of re-fetching — builds stay fast AND work
# offline. A cached copy older than this is treated as stale and re-fetched (best-effort).
_CACHE_DIR = os.path.join(_ROOT, "data", "funds", "portfolio_cache")
_CACHE_MAX_AGE_DAYS = 20

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

_MONTHS = ["January", "February", "March", "April", "May", "June",
           "July", "August", "September", "October", "November", "December"]

# section-header text -> asset class bucket
_ASSET_RULES = [
    ("equity", ("equity", "stock", "(a) listed", "preference shares")),
    ("debt", ("debt", "bond", "debenture", "ncd", "g-sec", "government sec", "gilt", "sdl", "pass through")),
    ("money_market", ("money market", "commercial paper", "certificate of deposit", "t-bill", "treasury bill", "cblo")),
    ("cash", ("treps", "tri-party", "repo", "net receivable", "net current asset", "cash", "margin")),
    ("mf_units", ("mutual fund", "units of", "reit", "invit")),
    ("derivative", ("derivative", "futures", "options", "hedging")),
]

# A valid ISIN is a 2-letter country code + 9 alphanumerics + 1 check digit (12 chars). We accept
# ANY country (not just India): equity funds like PPFAS hold foreign stocks (US…/MU…/LU… ISINs)
# that must still be counted in holdings + weights, else coverage silently undercounts.
_ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")

# explicit no-ISIN CASH line items that carry their own weight (TREPS / net receivables / repo).
# These are the bulk of a fund's non-security sleeve and must be captured so the asset mix sums
# to ~100; matched by name and never when the row is a subtotal ("total" guard) so we can't
# double-count a section subtotal.
_CASH_LINE_KEYS = ("treps", "tri-party", "reverse repo", "net receivable", "net current asset",
                   "net payable", "cblo", "cash & cash", "cash and cash", "clearing corporation",
                   "cash margin")


# --------------------------------------------------------------------------- AMC adapters
# Each adapter yields candidate (label, url) pairs newest-first; the first that returns a real
# workbook wins. The monthly file is named for the LAST day of the disclosed month, so we try
# the last few month-ends. Isolate all format/URL drift to this registry.
def _month_ends(n: int = 4):
    """The last `n` calendar month-end dates, newest first (today back)."""
    today = dt.date.today()
    first_this = today.replace(day=1)
    ends = []
    cur = first_this
    for _ in range(n):
        last_day = cur - dt.timedelta(days=1)          # last day of previous month
        ends.append(last_day)
        cur = last_day.replace(day=1)
    return ends


def _fy_label(d: dt.date) -> str:
    """Indian financial year label as Groww writes it, e.g. '2026 -2027' for any month in
    Apr-2026..Mar-2027 (note the exact ' -' spacing in their path)."""
    y = d.year if d.month >= 4 else d.year - 1
    return f"{y} -{y + 1}"


def _groww_urls():
    base = "https://assets-netstorage.growwmf.in/compliance_docs/Statutory Disclosure/Portfolio"
    for d in _month_ends(4):
        label = f"{_MONTHS[d.month - 1]} {d.day:02d}, {d.year}"
        yield (f"Groww {label}", f"{base}/{_fy_label(d)}/Monthly Portfolio- {label}.xlsx")


def _nippon_urls():
    base = "https://mf.nipponindiaim.com/InvestorServices/FactsheetsDocuments"
    for d in _month_ends(4):
        mon = _MONTHS[d.month - 1][:3].upper()
        yield (f"Nippon {mon}-{d.year}", f"{base}/MONTHLY-PORTFOLIO-{mon}-{d.year % 100:02d}.xls")


ADAPTERS = {
    "groww": {"name": "Groww Mutual Fund", "urls": _groww_urls},
    "nippon": {"name": "Nippon India Mutual Fund", "urls": _nippon_urls},
}


# --------------------------------------------------------------------------- registry (all-AMC)
# The two adapters above are hand-verified SEEDS. Every OTHER AMC is driven by a declarative
# URL-PATTERN registry on disk (data/funds/amc_feed_patterns.json), discovered by the local
# puppeteer harness (discover_amc_feeds.js) + the cloud discovery workflow. Add an AMC = add a
# spec there; NO code change. One spec:
#   { "amc": "<full name>", "landing": "<disclosure page>", "url_pattern": "<template|fixed url>",
#     "file_type": "xlsx|xls|zip", "method": "cdn|api|render-needed|waf-blocked|static",
#     "confidence": "high|medium|low", "notes": "..." }
# url_pattern placeholders, expanded per disclosure month-end (newest first):
#   {YYYY} {YY} {MM} {M} {Mon} {MON} {mon} {Month} {MONTH} {month} {DD} {D}
#   {FYS} {FYE} {FYS2} {FYE2}   (Indian financial year Apr->Mar: start/end year, 4- and 2-digit)
# A "static" method (or a url_pattern with no {placeholder}) is treated as one fixed URL.
# ★ THE WAF BRIDGE: for waf-blocked / render-needed feeds the datacenter build can't fetch even
# with the correct URL, so the harness (KV's residential IP) PRE-WARMS the very same byte-cache
# (_cache_path) and the build reads it OFFLINE — no doomed network call eats the time budget.
_PATTERNS_PATH = os.path.join(_ROOT, "data", "funds", "amc_feed_patterns.json")
_INDEX_PATH = os.path.join(_ROOT, "data", "funds", "amfi_amc_index.json")

# AMC name -> stable short key. The two seeds are pinned so they match the code adapters above.
_KEY_OVERRIDES = {
    "Groww Asset Management Limited": "groww",
    "Nippon Life India Asset Management Limited": "nippon",
}
_KEY_GENERIC = {"asset", "management", "managers", "manager", "mgmt", "company", "co", "ltd",
                "limited", "private", "pvt", "india", "indian", "investment", "investments",
                "amc", "of", "the", "and"}


def amc_key(name: str) -> str:
    """Deterministic short key for an AMC name — the SAME key used by the registry, the on-disk
    index, the byte-cache and the JS harness, so they never drift. (Mirror this logic in JS.)"""
    if name in _KEY_OVERRIDES:
        return _KEY_OVERRIDES[name]
    toks = re.sub(r"[^a-z0-9 ]+", " ", str(name or "").lower()).split()
    keep = [t for t in toks if t not in _KEY_GENERIC] or toks
    return "-".join(keep[:4])[:28].strip("-")


def _load_patterns() -> dict:
    """The on-disk AMC pattern specs, {key: spec}. Tolerates either {"amcs": {...}} or a bare map."""
    try:
        with open(_PATTERNS_PATH, encoding="utf-8") as f:
            d = json.load(f)
        if isinstance(d, dict):
            return d.get("amcs", d)
    except Exception:
        pass
    return {}


def _ordinal(n: int) -> str:
    """1->1st, 2->2nd, 3->3rd, 11->11th, 31->31st (lower-case suffix as SBI etc. write it)."""
    if 10 <= n % 100 <= 20:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suf}"


def _expand_pattern(pattern: str, d: dt.date) -> str:
    """Expand a url_pattern's placeholders for ONE disclosure month-end date `d`.
    Placeholders cover: the DATA month (the disclosed month-end), the PUBLISH month (data+1, since
    several AMCs file folders by upload month), the Indian financial year, and an ordinal day."""
    fys = d.year if d.month >= 4 else d.year - 1        # Indian FY runs Apr -> Mar
    fye = fys + 1
    mon = _MONTHS[d.month - 1]
    pub = (d.replace(day=1) + dt.timedelta(days=32)).replace(day=1)   # 1st of the month AFTER d
    pmon = _MONTHS[pub.month - 1]
    rep = {
        "{YYYY}": f"{d.year}", "{YY}": f"{d.year % 100:02d}",
        "{MM}": f"{d.month:02d}", "{M}": f"{d.month}",
        "{Mon}": mon[:3], "{MON}": mon[:3].upper(), "{mon}": mon[:3].lower(),
        "{Month}": mon, "{MONTH}": mon.upper(), "{month}": mon.lower(),
        "{DD}": f"{d.day:02d}", "{D}": f"{d.day}", "{DDord}": _ordinal(d.day),
        "{FYS}": f"{fys}", "{FYE}": f"{fye}", "{FYS2}": f"{fys % 100:02d}", "{FYE2}": f"{fye % 100:02d}",
        # PUBLISH month = data month + 1 (folders named for the upload month: HDFC, Bajaj, Helios…)
        "{PYYYY}": f"{pub.year}", "{PYY}": f"{pub.year % 100:02d}", "{PMM}": f"{pub.month:02d}",
        "{PMon}": pmon[:3], "{PMON}": pmon[:3].upper(), "{Pmonth}": pmon.lower(),
    }
    out = pattern
    for k, v in rep.items():
        out = out.replace(k, v)
    return out


# include / exclude vocab for the generic landing-page resolver: a monthly-portfolio file URL
# usually says "portfolio"/"monthly" and never "factsheet"/"nav"/"fortnight"/"annual"…
_RESOLVE_INC = ("portfolio", "monthly", "monthy", "mportfolio", "mp-", "_mp_", "59a", "reg59",
                "scheme-portfolio", "schemeportfolio", "holdings")
_RESOLVE_EXC = ("factsheet", "fact-sheet", "fact_sheet", "fortnight", "fortnightly", "weekly",
                "half-year", "halfyear", "half_year", "annual", "/nav", "nav-", "navhistory",
                "/sid", "/kim", "addendum", "notice", "riskometer", "risk-o-meter", "aum-",
                "ter-", "sip", "factsheets", "performance")
_FILE_RE = re.compile(r"""(?:https?://|/)[^\s'"<>()]+?\.(?:xlsx|xls|zip)\b""", re.I)
_JSON_URL_RE = re.compile(r'"[^"]*?\.(?:xlsx|xls|zip)[^"]*?"', re.I)


def _absolutize(u: str, base: str) -> str:
    try:
        from urllib.parse import urljoin
        return urljoin(base, u.replace(" ", "%20"))
    except Exception:
        return u


def _url_recency_score(u: str) -> tuple:
    """Rank a candidate file URL by the DATA as-on date it encodes — higher is newer.
    Returns (year, month, day); (0,0,0) if no date is parseable.

    CRITICAL: prefer the explicit AS-ON DATE spelled in the filename ('…as-on-31st-May-2026',
    'Monthly-Portfolio_May-2026') over loose token-maxing. Many AMCs file the disclosure under a
    PUBLISH-month folder one month AFTER the data month (Dec-2025 data lives under /2026/01/), so a
    naive max() of every '20xx' year token + every month name would score a Dec-2025 file as
    (2026,12) and wrongly rank it ABOVE the real-newest May-2026 file (published under /2026/06/).
    We parse the spelled-out 'Month YYYY' (with optional 'DDth' day) FIRST, and only fall back to
    token-maxing when the filename carries no spelled date.

    TRAP: a TARGET-MATURITY bond fund spells its maturity date in its OWN name
    ('nifty-g-sec-sep-2032-index-fund_may_2026_…'), which is NOT the disclosure date. The real
    as-on stamp is the one nearest the file extension, so we take the LAST spelled 'Month-YYYY'
    match (and ignore the '?sfvrsn=…' query) rather than the first — else a 2032 maturity in the
    slug would rank a current file as the newest of all."""
    # ★ URL-DECODE FIRST. Links lifted from a Next.js/Strapi hydration JSON arrive percent-encoded
    # (spaces -> %20), e.g. '…/ZNFTY%20-%20Monthly%20Portfolio%20May%202026.xlsx'. The spelled-date
    # regexes below key on a literal space/dash/underscore between 'Month' and 'YYYY', which '%20'
    # does NOT satisfy — so an encoded URL would mis-parse (the '%20' even injects a stray '20' that
    # the loose year token-max latches onto, scoring it (2020,…)). Decoding %20->space restores the
    # real as-on date and is harmless for already-decoded URLs.
    try:
        from urllib.parse import unquote
        u = unquote(u)
    except Exception:
        pass
    # Usually the date lives in the PATH and the query is a cache-buster (?sfvrsn=…) we must drop.
    # But some AMCs serve files via a PHP proxy where the workbook NAME (and its as-on date) lives
    # IN the query string — e.g. NJ MF: viewfile.php?file=NJ-MF-Monthly-Portfolio-NJABF-May-2026-….xlsx
    # (the bare path is just …/viewfile.php). So: keep the query when it carries a workbook filename,
    # else strip it so a stray ?sfvrsn=… can't be mis-read.
    head, _, query = u.partition("?")
    if re.search(r"\.(xlsx|xls|zip)\b", query, re.I) and not re.search(r"\.(xlsx|xls|zip)\b", head, re.I):
        low = u.lower()
    else:
        low = head.lower()
    months = {full.lower(): i for i, full in enumerate(_MONTHS, 1)}
    months.update({full[:3].lower(): i for i, full in enumerate(_MONTHS, 1)})
    mon_re = "|".join(sorted(months, key=len, reverse=True))
    # A disclosure as-on date can never be in the FUTURE. A target-maturity bond fund spells its
    # maturity (e.g. 'sep-2032') in the name; on an old hash-named file that maturity is the ONLY
    # date present, so it would otherwise score (2032,9) and rank as the "newest" of all. Clamp any
    # future (year,month) to (0,0,0) so a maturity token can never win a recency ranking.
    today = dt.date.today()
    cap = (today.year, today.month + 1) if today.month < 12 else (today.year + 1, 1)  # +1 mo slack
    def _ok(y, mth):
        return (y, mth, 0) if (y, mth) <= cap else (0, 0, 0)
    # 1) explicit as-on date in the filename: optional "DD(st/nd/rd/th)" then Month then 4-digit year.
    # Take the LAST such match — the disclosure stamp sits at the end of the name; any date in the
    # scheme slug (a target-maturity bond fund) comes earlier and must not win.
    ms = list(re.finditer(rf"(?:(\d{{1,2}})(?:st|nd|rd|th)?[\-_ ]+)?({mon_re})[\-_ ]+(20\d{{2}})", low))
    if ms:
        for m in reversed(ms):                          # newest-stamp-first, but skip future maturities
            sc = _ok(int(m.group(3)), months[m.group(2)])
            if sc != (0, 0, 0):
                day = int(m.group(1)) if m.group(1) else 0
                return (sc[0], sc[1], day)
    # 1b) 2-DIGIT year spelled after the month ('May_26', 'Apr_26', 'March26' = May/Apr/Mar 2026 —
    # Old Bridge and many houses abbreviate the year). Fires ONLY when no 4-digit 'Month YYYY' stamp
    # exists above, so it can never override a clean spelled date. The 2-digit value is read as a YEAR
    # (20-49 -> 2020-2049), never a day, so 'Portfolio_31' can't be mis-parsed; future months clamp via
    # _ok. Take the LAST in-range match (the stamp sits at the name's end, after any scheme-slug date).
    ms2 = list(re.finditer(rf"({mon_re})[\-_ ]*([2-4]\d)(?!\d)", low))
    if ms2:
        for m in reversed(ms2):
            sc = _ok(2000 + int(m.group(2)), months[m.group(1)])
            if sc != (0, 0, 0):
                return sc
    # 2) numeric YYYY-MM / YYYY_MM / YYYY/MM in the path (data, not publish, when only one present)
    m = re.search(r"(20\d{2})[\-_/](0[1-9]|1[0-2])\b", u)
    if m:
        return _ok(int(m.group(1)), int(m.group(2)))
    # 3) last-resort: loosest token-max (legacy behaviour) for filenames with no spelled date
    yr = 0
    ym = [int(y) for y in re.findall(r"(20\d{2})", low) if (int(y), 1) <= cap]   # ignore future years
    if ym:
        yr = max(ym)
    mo = 0
    for full, i in months.items():
        if full in low:
            mo = max(mo, i)
    m2 = re.search(r"(?:[_\-/](0[1-9]|1[0-2])[_\-/])", low)
    if m2:
        mo = max(mo, int(m2.group(1)))
    return _ok(yr, mo)


def _resolve_api(spec: dict, session, log=print):
    """DECLARATIVE API resolver for AMCs whose monthly-portfolio file index is a JSON endpoint
    behind a one-shot anti-CSRF NONCE — the file URLs are NOT in the static landing HTML and the
    endpoint refuses a plain GET (Navi's WordPress feed: POST /wp-json/nv/v1/documents with header
    WP-NONCE:<nonce scraped from the landing page>, body {financial_year, value:<MonthName>,
    category, type:"Monthly", order}). The generic landing-scrape can't drive this (no POST, no
    nonce), so a small spec block does it — staying registry-as-data: NO per-AMC code.

      spec['resolve']['api'] = {
        nonce_url   : page to GET for the nonce (defaults to resolve.landing / spec.landing),
        nonce_regex : capture-group-1 regex for the nonce (default WordPress '"nonce":"<hex>"'),
        nonce_header: header name carrying the nonce  (default "WP-NONCE"),
        url         : the API endpoint to POST (or GET) for the document index,
        method      : "POST" | "GET"                   (default "POST"),
        body        : form fields; values may use _expand_pattern placeholders ({Month},{FYS}-{FYE}…),
        data_path   : dotted path to the list of docs in the JSON  (default "data"),
        url_key     : key holding each doc's file URL   (default "url"),
        title_key   : key holding each doc's title/name (default "title"),
      }
    We expand the body for each recent month-end NEWEST-FIRST and stop at the first month that
    returns documents — so the LATEST disclosed month wins and we never serve a stale earlier month
    just because the API still lists it. Yields (title, url); pair with multi:true to fetch every
    per-scheme file."""
    rc = spec.get("resolve") or {}
    ac = rc.get("api") or {}
    api_url = ac.get("url")
    if not api_url or session is None:
        return
    nonce = None
    nu = ac.get("nonce_url") or rc.get("landing") or spec.get("landing")
    nonce_rx = ac.get("nonce_regex") or r'"nonce"\s*:\s*"([a-f0-9]+)"'
    if nu:
        try:
            rp = session.get(nu, timeout=(8, 25))
            mm = re.search(nonce_rx, rp.text or "")
            if mm:
                nonce = mm.group(1)
        except Exception:
            nonce = None
    headers = {}
    if nonce and ac.get("nonce_header", "WP-NONCE"):
        headers[ac.get("nonce_header", "WP-NONCE")] = nonce
    method = (ac.get("method") or "POST").upper()
    body_tpl = ac.get("body") or {}
    data_path = (ac.get("data_path") or "data").split(".")
    url_key, title_key = ac.get("url_key", "url"), ac.get("title_key", "title")
    emitted = False
    for d in _month_ends(int(ac.get("months") or 5)):       # newest month first
        body = {k: (_expand_pattern(v, d) if isinstance(v, str) else v) for k, v in body_tpl.items()}
        try:
            if method == "GET":
                r = session.get(api_url, params=body, headers=headers, timeout=(8, 25))
            else:
                r = session.post(api_url, data=body, headers=headers, timeout=(8, 25))
            j = r.json() if r.status_code == 200 else None
        except Exception:
            j = None
        if not isinstance(j, (dict, list)):
            continue
        node = j
        for seg in data_path:
            node = node.get(seg) if isinstance(node, dict) else None
        docs = node if isinstance(node, list) else (j if isinstance(j, list) else None)
        if not docs:
            continue
        n = 0
        for doc in docs:
            if not isinstance(doc, dict):
                continue
            url = doc.get(url_key)
            if not (isinstance(url, str) and re.search(r"\.(xlsx|xls|zip)\b", url, re.I)):
                continue
            title = doc.get(title_key) or spec.get("amc", "")
            try:                                            # the feed HTML-escapes the title (&#8211;)
                import html as _html
                title = _html.unescape(str(title))
            except Exception:
                title = str(title)
            yield (title, url.replace(" ", "%20"))
            n += 1
        if n:                                               # first month with files wins -> stop
            emitted = True
            break
    if not emitted:
        return


def _resolve_landing(spec: dict, session, log=print):
    """GENERIC resolver for AMCs whose monthly-portfolio URL is NOT date-derivable (random CMS
    hash / UUID / upload-timestamp / API slug). GET the landing page (static HTML or hydration
    JSON — most AMC pages embed the file links statically), collect every .xlsx/.xls/.zip URL,
    keep those that look like a monthly portfolio, and yield them newest-first. Returns nothing on
    failure (caller falls through to pattern / verified_url). Config is optional:
      spec['resolve'] = {landing?, include?[], exclude?[], pick?: 'newest'|'first'|'all',
                         only_latest?: bool}
    `landing` may be a STRING or a LIST of URLs (e.g. one API call per scheme-classification, as
    Invesco exposes — equity/fixed-income/hybrid/etc.); the file links from every URL are UNIONED.
    `only_latest`: after recency-sort keep ONLY files whose as-on date == the newest seen — drops the
    stale earlier-month files an API/CMS feed lists alongside the current month (Invesco's
    JanUrl..DecUrl all live in the same folder, so the include-vocab can't separate them by name)."""
    rc = spec.get("resolve") or {}
    landing = rc.get("landing") or spec.get("landing")
    if not landing or session is None:
        return
    landings = list(landing) if isinstance(landing, (list, tuple)) else [landing]
    inc = tuple(rc.get("include") or ()) + _RESOLVE_INC
    exc = tuple(rc.get("exclude") or ()) + _RESOLVE_EXC
    cands = set()
    for lp in landings:
        try:
            r = session.get(lp, timeout=(8, 25))
            if r.status_code != 200 or not r.text:
                continue
        except Exception:
            continue
        html = r.text
        for m in _FILE_RE.findall(html):
            cands.add(_absolutize(m.strip().strip('\\'), lp))
        for m in _JSON_URL_RE.findall(html):            # links inside Next.js/Strapi hydration JSON
            cands.add(_absolutize(m.strip('"').strip().strip('\\'), lp))
    out = []
    for absu in cands:
        low = absu.lower()
        if not any(k in low for k in inc):
            continue
        if any(k in low for k in exc):
            continue
        out.append(absu)
    if not out:
        return
    if rc.get("only_latest"):                           # keep just the newest as-on month present
        best = max(_url_recency_score(u) for u in out)
        if best > (0, 0, 0):                            # only when a real date was parseable
            out = [u for u in out if _url_recency_score(u) == best]
    pick = rc.get("pick", "newest")
    if pick == "first":
        ordered = out
    elif pick == "all":
        ordered = sorted(out, key=_url_recency_score, reverse=True)
    else:                                               # newest by date tokens, then by appearance
        ordered = sorted(out, key=_url_recency_score, reverse=True)
    cap = 200 if pick == "all" else 6                   # per-scheme houses (pick='all') need every link
    seen = set()
    for i, u in enumerate(ordered[:cap]):
        if u in seen:
            continue
        seen.add(u)
        yield (f"{spec.get('amc','')} resolved#{i}", u)


def _pattern_urls(spec: dict):
    """Yield (label, url) newest-first for a registry spec: expand its url_pattern over the last
    few month-ends, or yield one fixed URL for a static spec.

    ★ url_pattern may be a STRING or a LIST of templates. A LIST serves a house that splits one
    month across SEVERAL combined workbooks (e.g. LIC: a separate Equity and a Debt monthly file,
    both named for the same as-on month). We expand the WHOLE list for the newest month FIRST, then
    the whole list for the next month, etc. — so with multi:true + max_files=len(list) the engine
    grabs every class of the SAME (newest) month and never bleeds into the prior month."""
    pats = spec.get("url_pattern") or ""
    if not pats:
        return
    pat_list = [p for p in (pats if isinstance(pats, (list, tuple)) else [pats]) if p]
    name = spec.get("amc", "")
    static = spec.get("method") == "static"
    has_ph = lambda p: bool(re.search(r"\{[A-Za-z0-9]+\}", p))
    # ★ {slug} expansion — a per-scheme MULTI house (HSBC) whose CDN files are date-derivable EXCEPT
    # for a stable per-scheme slug in the path/filename. The spec carries spec['slugs'] = [list of
    # scheme slugs]; for a pattern containing {slug} we expand it for EVERY (month-end × slug). The
    # slugs are stable month-to-month, so the same list rolls forward to next month automatically and
    # the engine fetches each scheme straight from the fast CDN — no dependence on a flaky landing
    # page. Patterns WITHOUT {slug} are unaffected.
    slugs = [s for s in (spec.get("slugs") or []) if s]
    def _expand_slugs(pat: str, d: dt.date):
        base = _expand_pattern(pat, d)
        if "{slug}" not in pat:
            yield base
            return
        for sg in slugs:
            yield base.replace("{slug}", sg)
    for pat in pat_list:                                    # fixed/static templates emit once, no month loop
        if static or not has_ph(pat):
            yield (f"{name} (static)", pat)
    if static or not any(has_ph(p) for p in pat_list):
        return
    for d in _month_ends(5):                                # for each month-end, expand EVERY templated pattern
        lab = f"{_MONTHS[d.month - 1][:3]}-{d.year}"
        for pat in pat_list:
            if has_ph(pat):
                for u in _expand_slugs(pat, d):
                    yield (f"{name} {lab}", u)


def _candidates(spec: dict, session, offline: bool):
    """Yield (label, url) candidate workbook URLs for a spec, best strategy first:
      1) RESOLVE  — scrape the landing page for the live link (most robust; current-month-true),
      2) PATTERN  — expand a deterministic url_pattern over recent month-ends,
      3) VERIFIED — the last-known-good direct URL(s) the discovery actually downloaded.
    The first candidate whose bytes are a real workbook (PK/D0CF, or a zip we can open) wins.

    ★ verified_urls (LIST) — an EXPLICIT per-scheme URL list for a MULTI house whose file index is
    neither scrapable nor date-templatable (e.g. JM/Mirae/UTI: the landing is a React SPA and the
    file list comes from an AES-encrypted / JS-only API, and the per-scheme filenames are
    inconsistently spaced/punctuated so no url_pattern can produce them). The discovery harness
    decrypts that API, resolves every current-month per-scheme direct URL, and bakes the list here;
    the engine just yields each (cache-first fetch) and, with multi:true, concats all schemes. This
    is the encrypted-API analogue of the WAF byte-cache bridge — the harness does the enumeration
    the datacenter build can't, but the URLs stay plain GETs so the build needs no JS/crypto.
    Yielded BEFORE the single verified_url so a full per-scheme set always wins over the 1-file
    fallback."""
    gen = spec.get("_urls")                             # hand-verified legacy generator (groww/nippon seed)
    if callable(gen):
        try:
            yield from gen()
        except Exception:
            pass
    if spec.get("resolve") and not offline and session is not None:
        rc = spec.get("resolve") or {}
        if rc.get("api"):                                   # nonce-gated JSON document index (Navi) — try first
            try:
                yield from _resolve_api(spec, session)
            except Exception:
                pass
        try:
            yield from _resolve_landing(spec, session)
        except Exception:
            pass
    yield from _pattern_urls(spec)
    vus = spec.get("verified_urls")
    if isinstance(vus, (list, tuple)):
        nm = spec.get("amc", "")
        for i, u in enumerate(vus):
            if isinstance(u, str) and u:
                yield (f"{nm} verified#{i}", u)
    vu = spec.get("verified_url")
    if vu:
        yield (f"{spec.get('amc','')} verified", vu)


def _seed_specs() -> dict:
    """Fallback specs derived from the two hand-verified code adapters — used ONLY if the on-disk
    pattern registry is missing, so the Funds tab never regresses below Groww+Nippon."""
    out = {}
    for key in ("groww", "nippon"):
        a = ADAPTERS[key]
        out[key] = {"amc": a["name"], "method": "cdn", "file_type": "xlsx",
                    "_urls": a["urls"]}                 # legacy generator carried through verbatim
    return out


def _registry() -> dict:
    """The merged AMC fetch registry, keyed by amc_key. Each value is the SPEC dict plus a derived
    `offline_only` flag. Source of truth = the on-disk pattern registry (amc_feed_patterns.json),
    discovered by the workflow + the local harness; the two code adapters seed it if the file is
    absent. A spec is INCLUDED only if it has SOME way to fetch (resolve, url_pattern, verified_url,
    or a legacy generator)."""
    specs = _load_patterns() or _seed_specs()
    reg = {}
    for key, spec in specs.items():
        if not (spec.get("resolve") or spec.get("url_pattern") or spec.get("verified_url") or spec.get("_urls")):
            continue                                    # discovered-but-not-yet-fetchable (awaits harness)
        method = (spec.get("method") or "").lower()
        s = dict(spec)
        s.setdefault("amc", key)
        # WAF/render feeds are unreachable from a datacenter IP — read the harness-warmed cache only.
        s["offline_only"] = method in ("waf-blocked", "render-needed", "local-cache")
        reg[key] = s
    return reg


# --------------------------------------------------------------------------- fetch + open
def _session():
    if requests is None:
        return None
    s = requests.Session()
    s.headers.update({"User-Agent": UA, "Accept-Language": "en-US,en;q=0.9"})
    return s


def _cache_path(url: str) -> str:
    """Stable on-disk cache filename for a workbook URL (last 150 safe chars keep it unique+short)."""
    return os.path.join(_CACHE_DIR, re.sub(r"[^A-Za-z0-9._-]+", "_", url)[-150:] + ".bin")


def _is_workbook(b) -> bool:
    return bool(b) and b[:2] in (b"PK", b"\xd0\xcf")   # PK = xlsx zip; D0CF = legacy BIFF .xls


def _fetch(url: str, session, offline: bool = False,
           max_age_days: int = _CACHE_MAX_AGE_DAYS) -> bytes | None:
    """Workbook bytes for `url`, cache-first so a rebuild never needlessly re-downloads.
      - FRESH cache (< max_age_days old)  -> return it, no network.
      - offline=True                       -> cache-only (any age), never touch the network.
      - else fetch with a BOUNDED (connect, read) timeout; on success write the cache; on failure
        fall back to a STALE cache if one exists (stale holdings beat none — graceful-degrade)."""
    cp = _cache_path(url)
    try:
        if os.path.exists(cp):
            age_days = (time.time() - os.path.getmtime(cp)) / 86400.0
            if offline or age_days <= max_age_days:
                with open(cp, "rb") as f:
                    b = f.read()
                if _is_workbook(b):
                    return b
    except Exception:
        pass
    if offline or session is None:
        return None
    try:
        r = session.get(url, timeout=(8, 25))          # (connect, read) — bounded so a build can't hang
        if r.status_code == 200 and _is_workbook(r.content):
            try:
                os.makedirs(_CACHE_DIR, exist_ok=True)
                with open(cp, "wb") as f:
                    f.write(r.content)
            except Exception:
                pass
            return r.content
    except Exception:
        pass
    try:                                                # network failed — last resort: a stale cache
        if os.path.exists(cp):
            with open(cp, "rb") as f:
                b = f.read()
            if _is_workbook(b):
                return b
    except Exception:
        pass
    return None


def _open_workbook(content: bytes):
    """Open a portfolio workbook from bytes. The .xls/.xlsx extension lies on some AMC feeds, so
    sniff the magic: 'PK' -> OOXML or a ZIP-of-workbooks; legacy BIFF -> xlrd. Several AMCs ship a
    .zip that CONTAINS the real .xls/.xlsx (ICICI, ABSL, UTI, DSP) — an .xlsx is itself a zip, so we
    disambiguate by looking for .xls/.xlsx MEMBERS (an .xlsx never contains those) and recurse.
    Returns a list of (sheet_name, rows) across all member workbooks, or None."""
    if not content:
        return None
    head = content[:2]
    if head == b"PK":
        members, zf = [], None
        try:
            import zipfile
            zf = zipfile.ZipFile(io.BytesIO(content))
            members = [n for n in zf.namelist() if n.lower().endswith((".xls", ".xlsx"))]
        except Exception:
            members, zf = [], None
        if members and zf is not None:                  # ZIP-of-workbooks -> open each member, concat
            out = []
            for n in members:
                try:
                    sub = _open_workbook(zf.read(n))
                except Exception:
                    sub = None
                if not sub:
                    continue
                # The MEMBER FILENAME is usually the full scheme name (ICICI:
                # 'ICICI Prudential Balanced Advantage Fund.xlsx'); the inner sheet tab is a cryptic
                # code ('BAF'). Carry the filename stem through as the sheet's name hint.
                stem = re.sub(r"\.(xls|xlsx)$", "", os.path.basename(n), flags=re.I)
                out.extend((stem, rows) for _snm, rows in sub)
            return out or None
        if load_workbook is not None:                   # genuine .xlsx
            try:
                wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                out = []
                for nm in wb.sheetnames:
                    ws = wb[nm]
                    out.append((nm, [tuple(r) for r in ws.iter_rows(values_only=True)]))
                wb.close()
                return out
            except Exception:
                return None
        return None
    if head == b"\xd0\xcf":                             # legacy BIFF .xls -> xlrd
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            out = []
            for sh in wb.sheets():
                rows = [tuple(sh.row_values(i)) for i in range(sh.nrows)]
                out.append((sh.name, rows))
            return out
        except Exception:
            return None
    return None


# --------------------------------------------------------------------------- parse one sheet
def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _asset_class(section_text: str) -> str:
    t = section_text.lower()
    for cls, keys in _ASSET_RULES:
        if any(k in t for k in keys):
            return cls
    return "other"


def _to_float(v):
    try:
        if v is None or v == "":
            return None
        return float(str(v).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return None


def _is_pct_header(c: str) -> bool:
    """A '% to net assets' column under any of the dialects AMCs actually use: SEBI standard
    ('% to Net Assets'), ICICI ('% to Nav'), and assorted ('% of NAV', '% to net', 'weightage')."""
    return ("net asset" in c or "% to net" in c or "% of net" in c
            or ("nav" in c and "%" in c) or "weightage" in c or "% holding" in c
            or "% to aum" in c or "% of aum" in c or ("aum" in c and "%" in c))


def _find_header(rows):
    """Return (header_row_index, {field: col_index}) by matching header TEXT, or (None, {}).
    Tolerant of the many AMC header dialects: the column LABELS differ across houses (name =
    'Name of Instrument' | 'Company/Issuer/Instrument Name' | 'Name of the Security'; pct = '% to
    Net Assets' | '% to Nav'; value = 'Market/Fair Value' | 'Exposure/Market Value'), so we match
    each field by a set of synonyms, never by a fixed position."""
    for i, row in enumerate(rows[:30]):
        cells = [_norm(c).lower() for c in row]
        has_isin = any("isin" in c and len(c) <= 12 for c in cells)
        if not (has_isin and any(_is_pct_header(c) for c in cells)):
            continue
        cmap = {}
        for j, c in enumerate(cells):
            if not c:
                continue
            if "isin" in c and len(c) <= 12 and "isin" not in cmap:
                cmap["isin"] = j
            elif _is_pct_header(c) and "pct" not in cmap:
                cmap["pct"] = j
            elif ("name of" in c or "instrument" in c or "company" in c or "issuer" in c
                  or "security" in c or "name of the company" in c) and "name" not in cmap:
                cmap["name"] = j
            elif ("industry" in c or "rating" in c or "sector" in c) and "industry" not in cmap:
                cmap["industry"] = j
            elif (c == "quantity" or "quantity" in c or c == "qty" or "no. of" in c
                  or "no of shares" in c or "units" in c) and "quantity" not in cmap:
                cmap["quantity"] = j
            elif ("market value" in c or "fair value" in c or "market/fair" in c
                  or "exposure" in c) and "mktval" not in cmap:
                cmap["mktval"] = j
            elif "maturity" in c and "maturity" not in cmap:
                cmap["maturity"] = j
            elif ("put/call" in c or c.startswith("put") or "put / call" in c) and "putcall" not in cmap:
                cmap["putcall"] = j
        if "isin" in cmap and "pct" in cmap and "name" in cmap:
            return i, cmap
    return None, {}


# ---- scheme-name hygiene (audit 2026-06-24): the AMC XLSX often puts a verbose INVESTMENT-OBJECTIVE
# sentence, a SEBI disclaimer, a riskometer bullet, or a "PORTFOLIO STATEMENT OF <name> AS ON <date>"
# wrapper above the holdings table. The old "contains fund/index + word-count" score let those BEAT the
# real (short) scheme name, so ~147/1298 schemes showed objective/disclaimer text as their name. These
# helpers RECOVER the real name where one exists and DROP the sheet only when none does (verified by an
# adversarial audit to drop 0 real schemes). ----
_HOUSE_ONLY_RE = re.compile(r"^[\w.&'\- ]{2,40}\bmutual\s+fund$", re.I)   # bare "<House> Mutual Fund"
_BULLET_LEAD_RE = re.compile(r"^[\*•●▪∙·\-–—#§]+\s")
# "PORTFOLIO STATEMENT OF/FOR <name> AS ON <date>" wrapper (Edelweiss, Bharat Bond) -> <name>.
_STMT_WRAP_RE = re.compile(r"^\s*(?:monthly\s+)?portfolio\s+statement\s+(?:of|for)\s+(.*?)\s+as\s+on\b.*$", re.I)
_ASON_TAIL_RE = re.compile(r"\s+as\s+on\b.*$", re.I)
# objective tail after a SEPARATOR or an opening PAREN — recovers the real head:
#   "Nippon India MNC Fund (An open ended equity scheme …)" -> "Nippon India MNC Fund"
#   "360 ONE Balanced Hybrid Fund - An open ended …"        -> "360 ONE Balanced Hybrid Fund"
# The '(' in the class is what fixes the 14 parenthetical false-positives the audit caught.
_OBJECTIVE_TAIL_RE = re.compile(
    r"\s*[\(\-–—:,]\s*(?:an?\s+)?(?:open[\s-]?ended|close[d]?[\s-]?ended|interval)\b.*$", re.I)
# Phrases that only appear in an OBJECTIVE / disclaimer / riskometer line, never inside a real name.
_OBJECTIVE_PHRASES = (
    "an open ended", "an open-ended", "a open ended", "an openended", "a close ended", "a closed ended",
    "a close-ended", "a closed-ended", "an interval", "investors should", "investors are advised",
    "consult their", "consult your", "financial advis", "riskometer", "risk-o-meter",
    "suitable for investors", "is suitable for", "this product is suitable", "aims to provide",
    "aims to generate", "seeks to provide", "seeks to generate", "seeks to achieve", "endeavours to",
    "endeavors to", "to generate long", "to generate capital", "to generate income",
    "capital appreciation", "safety and liquidity", "liquidity and reasonable", "investing in",
    "investment in equity", "predominantly in", "by investing", "tracking the", "that tracks the",
    "replicating", "replicate the", "to replicate", "to track the", "to correspond to",
    "subject to tracking", "no assurance", "may or may not", "past performance", "please read",
    "scheme related documents", "before investing", "the scheme seeks", "the scheme aims",
    "the investment objective", "objective of the scheme", "objective of the fund", "the primary objective",
)


def _looks_like_objective(s: str) -> bool:
    """True if `s` reads like an objective / disclaimer / riskometer / bullet / bare house-name (i.e. NOT
    a real scheme). Conservative: a real name with Fund/ETF/Index/FOF/Plan or an '(Erstwhile/Formerly …)'
    alias is NOT flagged unless it ALSO carries an objective phrase."""
    if not s:
        return False
    low = s.lower().strip()
    if _BULLET_LEAD_RE.match(s):
        return True
    if _HOUSE_ONLY_RE.match(low):
        return True
    if any(p in low for p in _OBJECTIVE_PHRASES):
        return True
    if len(s.split()) >= 14 and (low.endswith(".") or "," in low) \
            and not any(t in low for t in (" fund", " etf", " fof", " plan", " index", " scheme")):
        return True
    return False


def _recover_scheme_name(s: str) -> str:
    """Recover the real scheme head from a polluted cell: decode HTML entities, un-wrap a
    'PORTFOLIO STATEMENT OF <name> AS ON <date>' wrapper, strip a trailing 'AS ON <date>', then strip a
    separator/paren-led objective tail. Never strips an '(Erstwhile/Formerly …)' alias (not open/close-ended)."""
    if not s:
        return s
    s = html.unescape(s).strip()
    m = _STMT_WRAP_RE.match(s)
    if m and m.group(1).strip():
        s = m.group(1).strip()
    s = _ASON_TAIL_RE.sub("", s).strip()
    head = _OBJECTIVE_TAIL_RE.sub("", s).strip()
    return head if (head and len(head) >= 4) else s


def parse_sheet(name: str, rows) -> dict | None:
    """Parse one scheme worksheet -> a holdings dict, or None if it isn't a portfolio sheet."""
    hdr_i, cmap = _find_header(rows)
    if hdr_i is None:
        return None

    # scheme name: prefer the value beside an explicit "SCHEME NAME :" label (Helios & others put
    # 'Helios Mutual Fund' on the first line and the REAL scheme name beside a 'SCHEME NAME :' cell);
    # else fall back to the first non-empty cell above the header. Strip a leading "CODE-" prefix.
    scheme = ""
    for r in rows[:hdr_i]:
        cells = [_norm(c) for c in r]
        for j, c in enumerate(cells):
            if "scheme name" in c.lower():
                for k in range(j + 1, len(cells)):
                    if cells[k]:
                        scheme = cells[k]
                        break
            if scheme:
                break
        if scheme:
            break
    # Collect EVERY non-empty cell above the header as a name candidate. The real scheme name is
    # NOT always the first one: HSBC puts the AMC's legal name 'HSBC Mutual Fund' on row 0 and the
    # actual scheme 'HSBC Nifty 50 Index Fund' on row 1, so a first-cell-only pick grabs the wrong
    # line. We score all candidates below and keep the most scheme-like.
    header_cands = []
    for r in rows[:hdr_i]:
        for c in r:
            cc = _norm(c)
            if cc and cc not in header_cands:
                header_cands.append(cc)
    if not scheme:
        scheme = header_cands[0] if header_cands else ""
    scheme = re.sub(r"^\s*scheme\s*name\s*[:\-]\s*", "", scheme, flags=re.I)   # drop a "SCHEME NAME:" label
    scheme = re.sub(r"^\s*scheme\s*[:\-]\s*", "", scheme, flags=re.I)          # drop a "SCHEME:" label (UTI stacked sheet)
    scheme = re.sub(r"^[A-Z0-9]{2,6}\s*-\s*", "", scheme)        # drop "IB01-" style code prefix
    scheme = re.sub(r"\s*\([^()]*open[ -]?end[^()]*\)\s*$", "", scheme, flags=re.I).strip()  # drop "(An open-ended …)" suffix

    # Pick the best scheme name between the in-sheet header text and the `name` hint (sheet tab or,
    # for a zip member, the member filename). Some AMCs put only the AMC name above the header
    # (ICICI: 'ICICI Prudential') and carry the real scheme name in the filename; others (Groww)
    # carry it in the sheet. Score by whether it reads like a scheme ('Fund'/'Scheme'/word count).
    name_hint = re.sub(r"^[A-Z0-9]{2,6}\s*-\s*", "", _norm(name))

    def _scheme_score(s: str) -> float:
        if not s:
            return -1.0
        s2 = s.lower()
        # "<AMC> Mutual Fund" is the house's legal name, never a scheme -> push it below any real
        # scheme candidate (e.g. the zip-member filename 'ICICI Prudential BHARAT 22 FOF').
        penalty = -5.0 if "mutual fund" in s2 else 0.0
        return penalty + (3.0 if "fund" in s2 else 0.0) \
            + (1.0 if any(t in s2 for t in ("scheme", "plan", "fof", "etf", "index")) else 0.0) \
            + 0.1 * min(len(s.split()), 6)
    # Best of: the label/first-cell pick, every header-area cell (cleaned of the open-ended suffix),
    # and the sheet-tab/zip-member name hint. This lets the AMC's legal 'Mutual Fund' line lose to
    # the true scheme line whichever row it sits on.
    def _clean(s: str) -> str:
        s = re.sub(r"^\s*scheme\s*name\s*[:\-]\s*", "", s, flags=re.I)        # "SCHEME NAME:" label
        s = re.sub(r"^\s*scheme\s*[:\-]\s*", "", s, flags=re.I)               # "SCHEME:" label (UTI stacked)
        s = re.sub(r"\s*\([^()]*open[ -]?end[^()]*\)\s*$", "", s, flags=re.I).strip()
        # drop a trailing disclosure-period suffix some API titles carry: "… 1st - 31st May 2026"
        s = re.sub(r"\s*\d{1,2}(?:st|nd|rd|th)?\s*[–—\-]\s*\d{1,2}(?:st|nd|rd|th)?\s+"
                   r"[A-Za-z]+\s+\d{4}\s*$", "", s, flags=re.I).strip()
        return s
    # Build the pool, RECOVER the real head (un-wrap "PORTFOLIO STATEMENT OF…", strip objective tails),
    # then DROP candidates that still read like an objective so they can never win the score. Keep an
    # unfiltered copy as fallback so an all-objective sheet doesn't end up empty before the guard.
    raw_pool = [_recover_scheme_name(_clean(scheme)), _recover_scheme_name(_clean(name_hint))] \
        + [_recover_scheme_name(_clean(c)) for c in header_cands]
    cand_pool = [c for c in raw_pool if c and not _looks_like_objective(c)]
    if cand_pool:
        scheme = max(cand_pool, key=_scheme_score)
    else:
        scheme = (max(raw_pool, key=_scheme_score) if raw_pool else scheme)
    # FINAL GUARD: if the chosen name STILL looks like an objective / disclaimer / bullet / bare house-name
    # (a sheet whose only above-header text is the objective paragraph), DROP the sheet rather than emit a
    # garbage "scheme". A real name with Fund/ETF/Index/FOF/Plan or an "(Erstwhile/Formerly …)" alias survives.
    if not scheme or _looks_like_objective(scheme):
        return None

    # as-on date from a "Portfolio as on …" line. AMCs write it as "as on" OR "as of" (HSBC:
    # "Portfolio Statement as of May 31, 2026"), so match both. The date may sit IN the label cell
    # ("… as on 31-May-2026") OR in an ADJACENT cell (Helios: label "PORTFOLIO STATEMENT AS ON :"
    # then a separate cell holding "2026-05-31 00:00:00") — try both.
    asof = None
    for r in rows[:hdr_i + 1]:
        cells = [_norm(c) for c in r]
        for j, c in enumerate(cells):
            m = re.search(r"as o[nf]\s*:?\s*([0-9A-Za-z,\-/ ]+)", c, re.I)
            if m and m.group(1).strip():
                asof = m.group(1).strip()
                break
            # Navi: "Monthly Portfolio Statement for the month ended 31 May 2026" (no 'as on/of').
            m2 = re.search(r"month ended\s*:?\s*([0-9]{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+\s+[0-9]{4})",
                           c, re.I)
            if m2 and m2.group(1).strip():
                asof = m2.group(1).strip()
                break
            # Zerodha: "MONTHLY PORTFOLIO STATEMENT OF ZERODHA NIFTY 50 INDEX FUND FOR MAY 2026"
            # — the as-on stamp is just "FOR <Month> <Year>" (full English month + 4-digit year), with
            # no day and no 'as on/of'. Anchor on the title verb 'PORTFOLIO … FOR' so a stray 'for'
            # elsewhere can't false-match, and require a real month name. The data month is the period;
            # we report it as 'Month YYYY' (the disclosure is as-of that month-end).
            m3 = re.search(r"portfolio\b.*?\bfor\s+([A-Za-z]+\s+[0-9]{4})", c, re.I)
            if m3:
                _mtok = m3.group(1).split()[0].lower()
                if _mtok[:3] in {mn[:3].lower() for mn in _MONTHS}:
                    asof = m3.group(1).strip()
                    break
            # Same Zerodha title but with NO 'for' ("…ZERODHA NIFTY 50 ETF MAY 2026"). Only on a
            # 'MONTHLY PORTFOLIO STATEMENT' line, take the LAST spelled '<Month> <Year>' (the as-on
            # stamp is at the end of the title; any month/year in the scheme slug comes earlier).
            if re.search(r"monthly\s+portfolio\s+statement", c, re.I):
                ms3 = re.findall(r"([A-Za-z]+)\s+([0-9]{4})", c)
                for mon, yr in reversed(ms3):
                    if mon[:3].lower() in {mn[:3].lower() for mn in _MONTHS}:
                        asof = f"{mon} {yr}"
                        break
                if asof:
                    break
            if re.search(r"as o[nf]", c, re.I):                 # label-only cell -> take the next non-empty cell
                for k in range(j + 1, len(cells)):
                    if cells[k]:
                        asof = re.sub(r"\s+00:00:00$", "", cells[k]).strip()
                        break
                if asof:
                    break
        if asof:
            break

    ci, cn, cd = cmap.get("isin"), cmap.get("name"), cmap.get("industry")
    cq, cv, cp = cmap.get("quantity"), cmap.get("mktval"), cmap.get("pct")
    cmat, cpc = cmap.get("maturity"), cmap.get("putcall")
    holdings, cur_class, grand_total_pct = [], "equity", None

    for row in rows[hdr_i + 1:]:
        isin = _norm(row[ci]) if ci is not None and ci < len(row) else ""
        nm = _norm(row[cn]) if cn is not None and cn < len(row) else ""
        pct = _to_float(row[cp]) if cp is not None and cp < len(row) else None
        low = (isin + " " + nm).lower()

        if "grand total" in low:
            grand_total_pct = pct
            break
        if not _ISIN_RE.match(isin):
            if nm:
                low_nm = nm.lower()
                val = _to_float(row[cv]) if cv is not None and cv < len(row) else None
                # capture explicit no-ISIN CASH line items (TREPS / net receivables) that carry
                # their own weight — guarded by "total" so a section subtotal can't slip in
                if pct is not None and "total" not in low_nm and any(c in low_nm for c in _CASH_LINE_KEYS):
                    holdings.append({"isin": "", "name": nm, "industry": "", "quantity": None,
                                     "mktval_lakh": val, "pct": pct, "asset_class": "cash"})
                # otherwise a named no-ISIN row is a SECTION header -> remember its asset class. Some
                # AMCs (ICICI) print a SUBTOTAL value beside the section label, so we classify by the
                # NAME regardless of whether the row carries a value (don't gate on val is None).
                else:
                    cls = _asset_class(nm)
                    if cls != "other":
                        cur_class = cls
            continue

        holdings.append({
            "isin": isin,
            "name": nm,
            "industry": _norm(row[cd]) if cd is not None and cd < len(row) else "",
            "quantity": _to_float(row[cq]) if cq is not None and cq < len(row) else None,
            "mktval_lakh": _to_float(row[cv]) if cv is not None and cv < len(row) else None,
            "pct": pct,
            "asset_class": cur_class,
            "maturity": _norm(row[cmat]) if cmat is not None and cmat < len(row) else "",
            "putcall": _norm(row[cpc]) if cpc is not None and cpc < len(row) else "",
        })

    if not holdings:
        return None
    return {"name": scheme, "asof": asof, "grand_total_pct": grand_total_pct, "holdings": holdings}


# --------------------------------------------------------------------------- normalise + aggregate
def _normalise_pct(rec: dict):
    """Scale the % column to PERCENT (0-100). Detect fraction-vs-percent from the grand total
    (≈1 -> fraction) else the holdings sum (≤2 -> fraction). Mutates each holding's 'pct'."""
    gt = rec.get("grand_total_pct")
    s = sum(h["pct"] for h in rec["holdings"] if h.get("pct") is not None)
    is_fraction = (gt is not None and gt <= 2.0) or (gt is None and 0.5 <= s <= 2.0)
    if is_fraction:
        for h in rec["holdings"]:
            if h.get("pct") is not None:
                h["pct"] = h["pct"] * 100.0


def _resolve_symbols(rec: dict, isin_map: dict):
    """Resolve each holding's ISIN to our UNCHANGING vst_id and its current NSE symbol through the SINGLE
    idmap resolver (lineage + collisions + check-digit all in one place); fall back to funds_nav's own
    ISIN map only when idmap can't resolve. Stamps both h['vst_id'] (the permanent join key) and
    h['symbol'] (the current display ticker)."""
    from . import idmap
    for h in rec["holdings"]:
        isin = h.get("isin") or ""
        vid = idmap.resolve_to_vid(isin) if idmap.is_valid_isin(isin) else None
        sym = None
        if vid:
            r = idmap.vid_record(vid)
            sym = r["nse_symbol"] if r else None
        if not sym and isin_map:
            sym = isin_map.get(isin)
        h["vst_id"] = vid
        h["symbol"] = sym


def _aggregate(rec: dict) -> dict:
    """Precompute the panel-ready aggregates from the parsed holdings."""
    H = rec["holdings"]
    tot = sum(h["pct"] for h in H if h.get("pct")) or 0.0

    asset, sector = {}, {}
    for h in H:
        p = h.get("pct") or 0.0
        asset[h["asset_class"]] = asset.get(h["asset_class"], 0.0) + p
        if h["asset_class"] == "equity" and h.get("industry"):
            sector[h["industry"]] = sector.get(h["industry"], 0.0) + p

    eq = sorted([h for h in H if h.get("pct") is not None], key=lambda x: x["pct"], reverse=True)
    top = [{"name": h["name"], "symbol": h.get("symbol"), "industry": h["industry"],
            "pct": round(h["pct"], 2), "asset_class": h["asset_class"]} for h in eq[:10]]

    # concentration + ISIN-resolution coverage (by weight)
    herf = sum((h["pct"] / 100.0) ** 2 for h in H if h.get("pct")) if tot else None
    resolved_w = sum(h["pct"] for h in H if h.get("pct") and h.get("symbol"))
    return {
        "asset_alloc": {k: round(v, 2) for k, v in sorted(asset.items(), key=lambda kv: -kv[1])},
        "sector_alloc": {k: round(v, 2) for k, v in sorted(sector.items(), key=lambda kv: -kv[1])[:15]},
        "top_holdings": top,
        "concentration": {"top10_pct": round(sum(h["pct"] for h in eq[:10]), 2),
                          "herfindahl": round(herf, 4) if herf is not None else None,
                          "n_holdings": len(H)},
        "coverage": {"pct_sum": round(tot, 1),
                     "isin_resolved_pct": round(resolved_w, 1),
                     "n_holdings": len(H),
                     # >105% of NAV = a HEDGED/gross-exposure fund (arbitrage / multi-asset:
                     # long equity + cash collateral + derivative legs all reported) — not an error
                     "gross_exposure": bool(tot > 105)},
    }


# --------------------------------------------------------------------------- scheme-code match
def _slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def _core_name(name: str) -> str:
    """Normalise a scheme name for matching across AMC sheet vs AMFI master (drop plan/option)."""
    n = name.lower()
    for w in ("direct", "regular", "plan", "growth", "idcw", "dividend", "payout",
              "reinvestment", "option", "fund", "-"):
        n = n.replace(w, " ")
    return re.sub(r"[^a-z0-9]+", "", n)


def _match_code(scheme_name: str, master: dict) -> str | None:
    """Best-effort map AMC-sheet scheme name -> AMFI scheme code via normalised core name."""
    if not master:
        return None
    target = _core_name(scheme_name)
    if not target:
        return None
    for code, info in master.items():
        if _core_name(info.get("name", "")) == target:
            return code
    # loosen: containment either way (handles 'Largecap' vs 'Large Cap')
    for code, info in master.items():
        cn = _core_name(info.get("name", ""))
        if cn and (cn in target or target in cn):
            return code
    return None


# --------------------------------------------------------------------------- stacked-sheet split
def _first_cell_text(row) -> str:
    """First non-empty cell of a row, as text (the marker/label usually sits in column A)."""
    for c in row:
        if c is not None and str(c).strip():
            return str(c)
    return ""


def _split_sheets_on_marker(sheets, marker_re):
    """Split worksheets that STACK many schemes in one sheet into one (name, rows) chunk per scheme.

    A scheme block starts at any row whose first non-empty cell matches `marker_re` (e.g.
    'SCHEME CODE002STARTS') and runs until the next such row. We also harvest a friendly scheme-name
    HINT from the block's 'SCHEME: <name>' line so parse_sheet labels the record correctly even
    though the block has no 'scheme name :' cell. Sheets with NO marker pass through unchanged, so a
    mixed file (one stacked SEBI sheet + other normal sheets) still works."""
    # 'SCHEME: <name>' (colon REQUIRED so this can't match the 'SCHEME CODE…STARTS' marker row,
    # which carries no colon).
    name_re = re.compile(r"^\s*scheme\s*:\s*(.+)$", re.I)
    out = []
    for nm, rows in sheets:
        marks = [i for i, r in enumerate(rows) if marker_re.search(_first_cell_text(r))]
        if not marks:                                   # not a stacked sheet -> leave as-is
            out.append((nm, rows))
            continue
        for k, st in enumerate(marks):
            en = marks[k + 1] if k + 1 < len(marks) else len(rows)
            block = rows[st:en]
            hint = nm
            for r in block[:6]:                         # 'SCHEME: UTI - Large Cap Fund' -> name hint
                txt = _first_cell_text(r)
                if marker_re.search(txt):               # never the marker row itself
                    continue
                mm = name_re.match(txt)
                if mm and mm.group(1).strip():
                    hint = mm.group(1).strip()
                    break
            out.append((hint, block))
    return out


# --------------------------------------------------------------------------- build
def build_amc(amc_key: str, session=None, isin_map=None, log=print, offline: bool = False,
              registry=None) -> list:
    """Fetch + parse ONE AMC's monthly portfolio(s) -> list of per-scheme records (parsed +
    aggregated). [] on any failure (graceful-degrade). `offline` = use the byte-cache only, never
    hit the network. A per-AMC time budget guarantees a dead feed can't stall the whole build.
    COMBINED feeds ship one workbook holding every scheme (stop at the first good file); MULTI
    feeds (spec['multi']=true) ship one file PER scheme, so we fetch every candidate and concat."""
    reg = registry if registry is not None else _registry()
    adp = reg.get(amc_key)
    if not adp:
        return []
    # WAF/render feeds are cache-only from a datacenter IP (the harness warms the cache elsewhere).
    amc_offline = offline or bool(adp.get("offline_only"))
    session = session or (None if amc_offline else _session())
    if session is None and not amc_offline:
        return []
    if session is not None and adp.get("insecure_tls"):
        # A few AMC CDNs serve an INCOMPLETE/self-signed TLS chain (e.g. LIC: licmf.com sends only
        # the leaf cert, so a stock trust store rejects it with SSLError). For ONLY those feeds —
        # opted in per-spec — disable cert verification so the build can still fetch them. Scope is
        # this one session; the URL/host is pinned by the spec, so this is not a blanket relaxation.
        try:
            session.verify = False
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        except Exception:
            pass
    multi = bool(adp.get("multi"))
    max_files = int(adp.get("max_files") or (120 if multi else 1))
    contents = []                                       # [(label, bytes)] — 1 for combined, many for multi
    used = None
    deadline = time.time() + (120 if multi else 45)     # hard wall: a dead feed can't bleed minutes
    for label, url in _candidates(adp, session, amc_offline):
        c = _fetch(url, session, offline=amc_offline)
        if c:
            contents.append((label, c))
            used = used or label
            if not multi or len(contents) >= max_files:
                break
        if time.time() > deadline:
            if not contents:
                log(f"[funds_portfolio] {amc_key}: fetch budget exhausted — skipping")
            break
        if not amc_offline:
            time.sleep(0.2)
    if not contents:
        log(f"[funds_portfolio] {amc_key}: no workbook reachable")
        return []

    # ★ STACKED-SHEET SPLIT (registry-as-data). A few AMCs (UTI, some Tata files) put EVERY scheme
    # in ONE worksheet, each scheme block delimited by a marker row ('SCHEME CODE002STARTS') and
    # carrying its OWN header + holdings (no per-scheme 'Grand Total'). parse_sheet reads one scheme
    # per sheet, so it would fuse all 80+ schemes into a single bogus record (4700 holdings,
    # coverage ~7700%). When the spec declares split_marker:<regex>, we cut each worksheet's rows
    # into one chunk per marker BEFORE parsing — so each scheme becomes its own record. No
    # AMC-specific code; any stacked-layout house works by adding the one spec field.
    split_marker = adp.get("split_marker")
    split_re = None
    if split_marker:
        try:
            split_re = re.compile(split_marker, re.I)
        except Exception:
            split_re = None

    recs, seen = [], set()
    _GENERIC_TAB = re.compile(r"^(sheet|tabelle|hoja|feuil)\s*\d*$", re.I)
    for _lab, content in contents:
        sheets = _open_workbook(content) or []
        if split_re is not None:                        # explode each worksheet at the scheme markers
            sheets = _split_sheets_on_marker(sheets, split_re)
        for nm, rows in sheets:
            # When the worksheet tab is a generic 'Sheet1' AND the file is one-scheme-per-file (the
            # scheme name lives in the API document TITLE, not the sheet), fall back to the candidate
            # label as the name hint so parse_sheet can recover the real scheme name (Navi).
            hint = nm
            if len(sheets) == 1 and (not _norm(nm) or _GENERIC_TAB.match(_norm(nm))):
                hint = _lab or nm
            try:
                rec = parse_sheet(hint, rows)
            except Exception:
                rec = None
            if not rec:
                continue
            sig = (rec["name"].lower().strip(), len(rec["holdings"]))
            if sig in seen:                             # de-dupe a scheme that appears in two files
                continue
            seen.add(sig)
            _normalise_pct(rec)
            _resolve_symbols(rec, isin_map or {})
            rec["agg"] = _aggregate(rec)
            rec["amc"] = adp.get("amc") or adp.get("name") or amc_key
            recs.append(rec)
    log(f"[funds_portfolio] {amc_key}: {len(recs)} schemes parsed from "
        f"{len(contents)} file(s) ('{used}')")
    return recs


def build_all(outdir: str = _OUT_DEFAULT, amcs=None, scheme_master: dict | None = None,
              isin_map: dict | None = None, log=print, offline: bool = False) -> dict:
    """Build every AMC's portfolios -> write one JSON per scheme into `outdir`, return the
    manifest {file_key: scheme_name}. file_key = AMFI code when matched, else a name slug.
    `offline` = cache-only (no network); if a run produces nothing, the existing per-scheme files
    in `outdir` are KEPT and re-manifested so a transient outage can't blank the Funds tab."""
    os.makedirs(outdir, exist_ok=True)
    if isin_map is None:
        try:
            from . import funds_nav
            isin_map = funds_nav.load_isin_map()
        except Exception:
            isin_map = {}
    if scheme_master is None:
        try:
            from . import funds_nav
            scheme_master = funds_nav.load_master() if hasattr(funds_nav, "load_master") else {}
        except Exception:
            scheme_master = {}

    session = None if offline else _session()
    registry = _registry()
    manifest, n = {}, 0
    for amc_key in (amcs or list(registry)):
        # Houses flagged needs_harness (WAF / JS-API / per-scheme-only-in-a-browser) can't be
        # COMPLETELY fetched from a datacenter IP — a partial cache would ship a misleadingly
        # truncated portfolio (e.g. 1 of UTI's 82 schemes). Skip them until KV's local browser
        # harness warms a complete cache and clears the flag. (Better absent than wrong.)
        if registry.get(amc_key, {}).get("needs_harness"):
            log(f"[funds_portfolio] {amc_key}: skipped (needs_harness — awaiting local browser harness)")
            continue
        for rec in build_amc(amc_key, session=session, isin_map=isin_map, log=log,
                             offline=offline, registry=registry):
            # FINAL name hygiene for EVERY adapter (some, e.g. Edelweiss, name a scheme from a
            # "PORTFOLIO STATEMENT OF <name> AS ON <date>" wrapper that bypasses parse_sheet's recovery) —
            # un-wrap once here so the wrapper can never leak into the key/slug or the display name.
            rec["name"] = _recover_scheme_name(rec.get("name") or "") or rec.get("name") or ""
            code = _match_code(rec["name"], scheme_master or {})
            # file key: prefer the globally-unique AMFI code; else namespace the slug by AMC so two
            # houses' "Liquid Fund" can't collide. Bound the length (a long scheme name would blow
            # past the Windows 260-char path limit) with a stable hash suffix.
            if code:
                key = code
            else:
                key = f"{amc_key}-{_slug(rec['name'])}"
                if len(key) > 90:
                    import hashlib
                    key = key[:80].rstrip("-") + "-" + hashlib.md5(
                        (amc_key + rec["name"]).encode("utf-8")).hexdigest()[:8]
            out = {
                "code": code, "key": key, "name": rec["name"], "amc": rec.get("amc"),
                "asof": rec.get("asof"), "source": "AMC monthly portfolio disclosure (SEBI Reg 59A)",
                "holdings": [{"name": h["name"], "symbol": h.get("symbol"), "isin": h["isin"],
                              "vst_id": h.get("vst_id"),
                              "industry": h["industry"], "pct": round(h["pct"], 3) if h.get("pct") is not None else None,
                              "asset_class": h["asset_class"], "quantity": h.get("quantity"),
                              "mktval_lakh": h.get("mktval_lakh"), "maturity": h.get("maturity") or None,
                              "putcall": h.get("putcall") or None}
                             for h in rec["holdings"]],
            }
            out.update(rec["agg"])
            with open(os.path.join(outdir, f"{key}.json"), "w", encoding="utf-8") as f:
                json.dump(out, f, ensure_ascii=False, allow_nan=False)
            manifest[key] = rec["name"]
            n += 1
    if n == 0:
        # produced nothing this run (all feeds unreachable + cache empty) — DON'T blank the tab:
        # re-manifest whatever per-scheme files already exist on disk so the last good data persists.
        for fn in os.listdir(outdir):
            if fn.endswith(".json"):
                try:
                    with open(os.path.join(outdir, fn), encoding="utf-8") as f:
                        prev = json.load(f)
                    manifest[prev.get("key") or fn[:-5]] = prev.get("name") or fn[:-5]
                except Exception:
                    pass
        log(f"[funds_portfolio] no fresh data — kept {len(manifest)} existing scheme files in {outdir}")
    else:
        # remove ORPHANS — stale per-scheme files from a prior build whose key is no longer in the
        # manifest (e.g. slugs that changed after the name-hygiene fix), so the cockpit picker / bridge
        # never surface a stale duplicate. Only when fresh data was produced (n>0), never on a no-op run.
        keep = set(manifest)
        n_orphan = 0
        for fn in os.listdir(outdir):
            if fn.endswith(".json") and fn[:-5] not in keep:
                try:
                    os.remove(os.path.join(outdir, fn)); n_orphan += 1
                except OSError:
                    pass
        log(f"[funds_portfolio] wrote {n} scheme holdings files -> {outdir}"
            + (f" (removed {n_orphan} orphans)" if n_orphan else ""))
    return manifest


if __name__ == "__main__":
    import sys
    which = sys.argv[1:] or ["groww"]
    sess = _session()
    imap = {}
    try:
        from vistas import funds_nav
        imap = funds_nav.load_isin_map()
    except Exception as e:
        print("isin_map load failed:", e)
    for k in which:
        rs = build_amc(k, session=sess, isin_map=imap, log=lambda x: print(x, flush=True))
        for r in rs[:3]:
            a = r["agg"]
            print(f"\n=== {r['name']}  (as on {r['asof']}) ===")
            print("  asset:", a["asset_alloc"])
            print("  top3 :", [(h['name'][:24], h['symbol'], h['pct']) for h in a['top_holdings'][:3]])
            print("  cover: pct_sum={pct_sum}  isin_resolved={isin_resolved_pct}%  n={n_holdings}".format(**a["coverage"]))
