"""vistas/amc_firm.py — the Digital-AMC FIRM OS (multi-AMC paper-trading books).

Turns the Digital-AMC *reasoning demo* into a firm that actually HOLDS BOOKS, trades within
liquidity, marks to market every day, and keeps an auditable record it learns from. It recreates
real Indian fund houses as VIRTUAL firms: each real scheme (at its real latest AUM) gets a paper
book, run by our analyst/FM/CIO agents, and is scored against BOTH its benchmark AND the real
scheme's actual track (we hold the real NAV + attribution — a built-in scorecard).

Design (AGENTIC_AMC.md, Phase 1) — the engine/judgment split keeps it honest + tractable:
  • This module = the DETERMINISTIC firm OS: registry, books, blotter, mark-to-market daily
    fact sheet, liquidity caps, scorecards. Every NUMBER is produced here, reproducibly.
  • The historical REPLAY is driven by a rules-based FM constructor (build_rules in #68) — NOT
    the LLM committee (running LLMs over years x dozens of schemes is cost-prohibitive). The
    LLM analyst/FM/CIO agents take over LIVE-FORWARD, adding judgment on top of the rules.

STATE (git-tracked = the audit trail) under amc_book/ :
  _registry.json                          AMCs - schemes - real AUM - mandate - benchmark
  <AMC>/<SCHEME>/book.json                live positions + cash + NAV + play-type tags
                /blotter.jsonl            every trade, append-only (rationale + pitch/decision link)
                /prereg.jsonl             every thesis + its falsifier, before the outcome
                /daily/<YYYY-MM>.json      daily mark-to-market fact sheets (the CITI column schema)
                /scorecard.json           IC - IR=IC*sqrt(BR)*TC vs benchmark AND vs the real scheme
  _firm/{pitches,decisions,rulings}.jsonl + lessons/<agent>.md

DAILY FACT-SHEET schema (mirrors ABSL's CITI_BIRLA_DAILY_EQUITY_FACT_SHEET.xlsx):
  sr, isin, name, sector, bbg, qty, avg_cost, book_cost, mkt_price, mkt_value,
  prev_price, prev_value, pct_change, wtd_contribution, pct_assets   (+ sector subtotals,
  + footer: equity, cash, nav, day_return, since_inception)  — also rendered to .xlsx for audit.

CONVENTIONS (so every number is reproducible — KV reporting rule):
  AUM        = Σ mktval_lakh / 100 over the scheme's latest disclosed holdings (₹cr). The fund's
               real latest size; seeds the book as 100% CASH (per decision: agents deploy it).
  price      = the terminal's clean adjusted total-return close (vistas.stocks.load), so the paper
               NAV is a true total-return series. Marked daily even on no-trade days.
  liquidity  = a position is capped at min(mandate cap, LIQ_DAYS x trailing median daily turnover);
               a single day's trade ≤ TRADE_ADV_FRAC of that day's turnover (a big book takes days
               to build). Turnover from bhav OHLCV (as-of date); current median as a fallback.
  play_type  = structural (core long-term compounder) / cyclical (medium-term rotation tilt) /
               tactical (short-term catalyst). Tags every holding so the book is legible by horizon.
"""

import os
import json
import datetime as _dt

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
BOOK_DIR = os.path.join(_ROOT, "amc_book")

LIQ_DAYS = 20            # max position ≤ this many days of median daily turnover
TRADE_ADV_FRAC = 0.15   # a single day's trade ≤ this fraction of the day's turnover


# ───────────────────────────────────────────────────────── funds data location
def _funds_dirs():
    """Prefer the freshly-built site; fall back to the stable published copy. Returns
    (attribution_dir, portfolio_dir) or (None, None)."""
    for base in (os.path.join(_ROOT, "output", "terminal_site", "data"),
                 os.path.join(_ROOT, "_pages", "terminal", "data")):
        a = os.path.join(base, "funds_attribution")
        p = os.path.join(base, "funds_portfolio")
        if os.path.isdir(a) and os.path.isdir(p):
            return a, p
    return None, None


def _f(x, d=0.0):
    try:
        v = float(x)
        return v if v == v else d
    except Exception:
        return d


# ───────────────────────────────────────────────────────── mandates (SEBI category → rules)
# cap_floor: which AMFI market-cap buckets the mandate must live in (large=top100, mid=101-250,
# small=251+). equity_min: floor on equity %. max_pos / max_sector: single-name / sector caps.
# n_lo/n_hi: target holdings count band. turn: annual turnover budget (a soft style guide).
_M = lambda **k: k
MANDATES = {
    "Large Cap Fund":        _M(buckets=["large"], equity_min=0.80, max_pos=0.10, max_sector=0.35, n_lo=35, n_hi=55, turn=0.4),
    "Large & Mid Cap Fund":  _M(buckets=["large", "mid"], equity_min=0.70, max_pos=0.08, max_sector=0.35, n_lo=45, n_hi=70, turn=0.5),
    "Mid Cap Fund":          _M(buckets=["mid"], equity_min=0.65, max_pos=0.07, max_sector=0.35, n_lo=45, n_hi=70, turn=0.6),
    "Small Cap Fund":        _M(buckets=["small"], equity_min=0.65, max_pos=0.05, max_sector=0.30, n_lo=55, n_hi=90, turn=0.5),
    "Multi Cap Fund":        _M(buckets=["large", "mid", "small"], equity_min=0.75, max_pos=0.08, max_sector=0.35, n_lo=50, n_hi=80, turn=0.5),
    "Flexi Cap Fund":        _M(buckets=["large", "mid", "small"], equity_min=0.65, max_pos=0.09, max_sector=0.35, n_lo=40, n_hi=70, turn=0.5),
    "Focused Fund":          _M(buckets=["large", "mid", "small"], equity_min=0.65, max_pos=0.10, max_sector=0.40, n_lo=20, n_hi=30, turn=0.4),
    "ELSS":                  _M(buckets=["large", "mid", "small"], equity_min=0.80, max_pos=0.09, max_sector=0.35, n_lo=40, n_hi=65, turn=0.4),
    "Value Fund":            _M(buckets=["large", "mid", "small"], equity_min=0.65, max_pos=0.09, max_sector=0.35, n_lo=40, n_hi=70, turn=0.4, tilt="value"),
    "Contra Fund":           _M(buckets=["large", "mid", "small"], equity_min=0.65, max_pos=0.09, max_sector=0.35, n_lo=40, n_hi=70, turn=0.4, tilt="value"),
    "Dividend Yield Fund":   _M(buckets=["large", "mid", "small"], equity_min=0.65, max_pos=0.09, max_sector=0.35, n_lo=35, n_hi=60, turn=0.4, tilt="yield"),
    "Sectoral / Thematic":   _M(buckets=["large", "mid", "small"], equity_min=0.80, max_pos=0.12, max_sector=1.00, n_lo=25, n_hi=50, turn=0.5, thematic=True),
    "Aggressive Hybrid Fund": _M(buckets=["large", "mid"], equity_min=0.65, equity_max=0.80, max_pos=0.09, max_sector=0.35, n_lo=35, n_hi=60, turn=0.4),
    "Dynamic Asset Allocation or Balanced Advantage": _M(buckets=["large", "mid"], equity_min=0.30, equity_max=0.90, max_pos=0.08, max_sector=0.35, n_lo=35, n_hi=60, turn=0.6, daa=True),
    "Multi Asset Allocation": _M(buckets=["large", "mid"], equity_min=0.40, equity_max=0.80, max_pos=0.08, max_sector=0.35, n_lo=30, n_hi=55, turn=0.5),
    "Equity Savings":        _M(buckets=["large", "mid"], equity_min=0.20, equity_max=0.50, max_pos=0.07, max_sector=0.35, n_lo=30, n_hi=55, turn=0.6),
}
_DEFAULT_MANDATE = _M(buckets=["large", "mid", "small"], equity_min=0.65, max_pos=0.09, max_sector=0.35, n_lo=40, n_hi=70, turn=0.5)


def mandate_for(category):
    return dict(MANDATES.get(str(category or "").strip(), _DEFAULT_MANDATE))


# ───────────────────────────────────────────────────────── registry (real AMCs → schemes)
def _read_json(p):
    try:
        return json.loads(open(p, encoding="utf-8").read())
    except Exception:
        return None


def _vset(holds):
    return frozenset(h.get("vst_id") for h in (holds or []) if h.get("vst_id"))


def _portfolio_list(pdir):
    """Every disclosed scheme portfolio as {vs(vst_id set of equity holdings), aum_cr, asof,
    n_equity, holdings:[...], name, amc}. Joined to attribution by HOLDINGS FINGERPRINT (vst_id
    set overlap) — robust to name/plan/format differences (the funds_bridge method), since the
    portfolio files carry no navindia_code (only a slug key)."""
    import glob
    out = []
    for fp in glob.glob(os.path.join(pdir, "*.json")):
        if os.path.basename(fp).startswith("_"):
            continue
        d = _read_json(fp)
        if not d:
            continue
        hold = d.get("holdings") or []
        if not isinstance(hold, list):
            continue
        eq = [h for h in hold if str(h.get("asset_class")) == "equity"]
        vs = _vset(eq)
        if not vs:
            continue
        aum = sum(_f(h.get("mktval_lakh")) for h in hold) / 100.0     # lakh → ₹cr
        out.append({"vs": vs, "aum_cr": round(aum, 1), "asof": d.get("asof"),
                    "n_equity": len(eq), "holdings": eq, "name": d.get("name"), "amc": d.get("amc")})
    return out


def _best_portfolio(av, plist, thresh=0.5):
    """Best holdings-fingerprint match (Jaccard of vst_id sets) for an attribution scheme's
    equity holdings `av`, or None below `thresh`."""
    best, bs = None, thresh
    for p in plist:
        u = len(av | p["vs"])
        if not u:
            continue
        j = len(av & p["vs"]) / u
        if j >= bs:
            bs, best = j, p
    return (best, round(bs, 3)) if best else (None, 0.0)


def registry(amcs=None, categories=None, min_aum_cr=0.0):
    """Build the AMC→scheme registry from the real funds data.

    amcs/categories: optional allow-lists (substring match on AMC, exact on SEBI category).
    Returns {amc: {scheme_code: {scheme, amc, category, benchmark, aum_cr, asof, mandate,
    real_holdings, real_cagr, real_bench_cagr, real_ir, real_verdict}}}.
    """
    import glob
    adir, pdir = _funds_dirs()
    if not adir:
        return {}
    plist = _portfolio_list(pdir)
    out = {}
    for fp in glob.glob(os.path.join(adir, "*.json")):
        if os.path.basename(fp).startswith("_"):
            continue
        a = _read_json(fp)
        if not a:
            continue
        amc = (a.get("amc") or "").strip()
        cat = a.get("sebi_category")
        if amcs and not any(s.lower() in amc.lower() for s in amcs):
            continue
        if categories and cat not in categories:
            continue
        code = str(a.get("navindia_code"))
        # holdings fingerprint = the scheme's own equity vst_id set (full book, from crowd_flow)
        av = _vset(((a.get("crowd_flow") or {}).get("equity_holdings")) or
                   ((a.get("portfolio") or {}).get("top_holdings")))
        pf, score = (_best_portfolio(av, plist) if av else (None, 0.0))
        aum = _f((pf or {}).get("aum_cr"))
        if aum < min_aum_cr:
            continue
        out.setdefault(amc, {})[code] = {
            "scheme": a.get("scheme_name"), "amc": amc, "category": cat,
            "benchmark": a.get("benchmark"), "code": code,
            "aum_cr": aum, "asof": (pf or {}).get("asof"), "n_equity": (pf or {}).get("n_equity"),
            "match_score": score, "equity_pct_fund": (a.get("portfolio") or {}).get("equity_pct_fund"),
            "mandate": mandate_for(cat),
            "real_holdings": (pf or {}).get("holdings") or [],
            "real_cagr": a.get("cagr_paper"), "real_bench_cagr": a.get("cagr_bench"),
            "real_ir": a.get("info_ratio"), "real_verdict": a.get("verdict"),
        }
    return out


# ───────────────────────────────────────────────────────── prices / liquidity
_PX = {"df": None}


def _prices():
    """Cached Date×Symbol adjusted total-return price panel."""
    if _PX["df"] is None:
        from . import stocks as _st
        _PX["df"] = _st.load()
    return _PX["df"]


def price_asof(symbol, asof=None):
    """Latest adjusted close for `symbol` on/before `asof` ('YYYY-MM-DD' or None=latest), or None."""
    df = _prices()
    if df is None or symbol not in df.columns:
        return None
    s = df[symbol].dropna()
    if asof:
        s = s[s.index <= asof]
    if not len(s):
        return None
    return float(s.iloc[-1])


_TURN = {"df": None}


def _turnover_panel():
    """Cached {symbol: (sorted dates, turnover ₹cr array)} from bhav OHLCV. Heavy cold-read;
    guarded — returns {} on failure (then liquidity falls back to the quant snapshot)."""
    if _TURN["df"] is None:
        try:
            from . import bhav_prices as _bp
            d = _bp.load_ohlcv(adjusted=False)
            g = {}
            for sym, sub in d.groupby("sym"):
                sub = sub.sort_values("date")
                g[str(sym)] = (list(sub["date"].astype(str)), [(_f(t) / 1e7) for t in sub["turnover"]])  # ₹ → ₹cr
            _TURN["df"] = g
        except Exception:
            _TURN["df"] = {}
    return _TURN["df"]


def median_turnover_cr(symbol, asof=None, window=21):
    """Trailing-`window` median daily traded value (₹cr) for `symbol` on/before `asof`."""
    g = _turnover_panel().get(str(symbol))
    if not g:
        return None
    dates, vals = g
    if asof:
        n = sum(1 for d in dates if d <= str(asof))
    else:
        n = len(dates)
    seg = [v for v in vals[max(0, n - window):n] if v and v > 0]
    if not seg:
        return None
    seg.sort()
    return seg[len(seg) // 2]


def liquidity_cap_cr(symbol, aum_cr, asof=None):
    """Max ₹cr a book may hold in `symbol` = LIQ_DAYS × trailing median daily turnover.
    Returns None if turnover unknown (caller treats as 'no extra cap beyond the mandate')."""
    mt = median_turnover_cr(symbol, asof)
    if mt is None:
        return None
    return LIQ_DAYS * mt


# ───────────────────────────────────────────────────────── book + daily fact sheet
def new_book(reg_entry):
    """A fresh 100%-cash book seeded with the scheme's real latest AUM."""
    return {
        "scheme": reg_entry["scheme"], "amc": reg_entry["amc"], "code": reg_entry["code"],
        "category": reg_entry["category"], "benchmark": reg_entry["benchmark"],
        "aum0_cr": reg_entry["aum_cr"], "cash_cr": reg_entry["aum_cr"],
        "inception": None, "asof": None,
        "positions": {},   # symbol -> {isin, name, sector, qty, avg_cost, play_type, entry_date, thesis_ref}
        "nav0": 100.0,
    }


def _sector_of(pos):
    return pos.get("sector") or "Unclassified"


def fact_sheet(book, asof, prev_asof=None):
    """Mark the book to market on `asof` → the daily fact sheet (CITI schema). Pure: reads only
    the price panel. Returns {header, rows, sectors, footer}. qty is in SHARES; ₹ values in CRORE.

    avg_cost / mkt_price are PER-SHARE (₹); book_cost / mkt_value are ₹cr (= qty·price/1e7).
    wtd_contribution = pct_assets · pct_change/100 (the holding's points-contribution to the day).
    """
    rows = []
    eq_val = 0.0
    pos = book.get("positions", {})
    for sym, p in pos.items():
        qty = _f(p.get("qty"))
        if qty <= 0:
            continue
        mp = price_asof(sym, asof)
        pp = price_asof(sym, prev_asof) if prev_asof else None
        if mp is None:
            continue
        mv = qty * mp / 1e7        # ₹ → ₹cr
        bc = qty * _f(p.get("avg_cost")) / 1e7
        pv = (qty * pp / 1e7) if pp is not None else None
        pct_chg = ((mp / pp - 1) * 100.0) if (pp and pp > 0) else None
        eq_val += mv
        rows.append({"isin": p.get("isin"), "name": p.get("name") or sym, "sym": sym,
                     "sector": _sector_of(p), "play_type": p.get("play_type"),
                     "qty": round(qty), "avg_cost": round(_f(p.get("avg_cost")), 4),
                     "book_cost": round(bc, 4), "mkt_price": round(mp, 4), "mkt_value": round(mv, 4),
                     "prev_price": (round(pp, 4) if pp is not None else None),
                     "prev_value": (round(pv, 4) if pv is not None else None),
                     "pct_change": (round(pct_chg, 4) if pct_chg is not None else None)})
    cash = _f(book.get("cash_cr"))
    total = eq_val + cash
    # % assets + weighted contribution (needs total)
    for r in rows:
        r["pct_assets"] = round(100.0 * r["mkt_value"] / total, 4) if total > 0 else None
        r["wtd_contribution"] = (round(r["pct_assets"] * r["pct_change"] / 100.0, 5)
                                 if (r["pct_assets"] is not None and r["pct_change"] is not None) else None)
    rows.sort(key=lambda r: (r["sector"], -r["mkt_value"]))
    for i, r in enumerate(rows, 1):
        r["sr"] = i
    # sector subtotals
    sect = {}
    for r in rows:
        s = sect.setdefault(r["sector"], {"sector": r["sector"], "mkt_value": 0.0, "pct_assets": 0.0})
        s["mkt_value"] += r["mkt_value"]
        s["pct_assets"] += (r["pct_assets"] or 0.0)
    sectors = sorted(sect.values(), key=lambda s: -s["mkt_value"])
    day_ret = None
    prev_total = sum((r["prev_value"] or 0.0) for r in rows) + cash
    if prev_asof and prev_total > 0:
        day_ret = round((total / prev_total - 1) * 100.0, 4)
    return {
        "header": {"scheme": book.get("scheme"), "amc": book.get("amc"),
                   "asof": str(asof), "aum_cr": round(total, 2)},
        "rows": rows, "sectors": [{"sector": s["sector"], "pct_assets": round(s["pct_assets"], 3),
                                   "mkt_value": round(s["mkt_value"], 3)} for s in sectors],
        "footer": {"equity_cr": round(eq_val, 3), "cash_cr": round(cash, 3),
                   "total_cr": round(total, 3), "cash_pct": round(100.0 * cash / total, 2) if total > 0 else None,
                   "n_holdings": len(rows), "day_return_pct": day_ret},
    }


# ───────────────────────────────────────────────────────── state IO (git-tracked audit trail)
def _safe(s):
    return "".join(c if (c.isalnum() or c in " -_") else "_" for c in str(s)).strip().replace("  ", " ")


def scheme_dir(amc, scheme):
    d = os.path.join(BOOK_DIR, _safe(amc), _safe(scheme))
    os.makedirs(d, exist_ok=True)
    return d


def save_book(book):
    d = scheme_dir(book["amc"], book["scheme"])
    with open(os.path.join(d, "book.json"), "w", encoding="utf-8") as f:
        json.dump(book, f, indent=1, default=str)
    return d


def append_blotter(amc, scheme, trade):
    d = scheme_dir(amc, scheme)
    with open(os.path.join(d, "blotter.jsonl"), "a", encoding="utf-8") as f:
        f.write(json.dumps(trade, default=str) + "\n")


def save_daily(amc, scheme, sheet):
    """Bucket daily fact sheets by month so the audit trail stays navigable."""
    d = os.path.join(scheme_dir(amc, scheme), "daily")
    os.makedirs(d, exist_ok=True)
    ym = str(sheet["header"]["asof"])[:7]
    path = os.path.join(d, f"{ym}.json")
    book = _read_json(path) or {}
    book[str(sheet["header"]["asof"])] = sheet
    with open(path, "w", encoding="utf-8") as f:
        json.dump(book, f, default=str)


# ───────────────────────────────────────────────────────── rules-FM constructor (v0; seed of #68)
# The DETERMINISTIC FM that deploys a 100%-cash book into a real, constrained portfolio — no LLM,
# no look-ahead, reproducible. v0 deploys into the scheme's own disclosed-holdings universe (the
# investable set we can price), scored by our VALIDATED edge (ARM = analyst-revision momentum, the
# 0-100 percentile; IC ~0.03-0.045 on our data), then water-fills weights under the mandate's
# single-name + sector caps AND a hard liquidity cap, tagging each name by revision-cycle play-type.
# This is the engine the historical REPLAY (#68) will walk forward; the LLM agents refine it live.
_ARM = {"raw": None}


def _arm_raw():
    if _ARM["raw"] is None:
        try:
            from . import arm as _a
            _ARM["raw"] = _a.load_raw()
        except Exception:
            _ARM["raw"] = {}
    return _ARM["raw"]


def current_arm(isin, asof=None, stale_days=400):
    """Latest ARM_100_REG (0-100 analyst-revision percentile) for `isin` on/before `asof`, or None
    if absent or STALE (last revision > stale_days before asof — a dead/uncovered name we won't tilt
    on). Licensed input used only to SCORE selection; the book stores aggregates, not per-stock ARM."""
    rec = _arm_raw().get(isin)
    if not rec:
        return None
    series = (rec.get("mnem") or {}).get("ARM_100_REG")
    if not series:
        return None
    pts = [(d, v) for d, v in series if (asof is None or str(d) <= str(asof)) and v == v]
    if not pts:
        return None
    last_d, last_v = pts[-1]
    if asof:
        try:
            gap = (_dt.date.fromisoformat(str(asof)) - _dt.date.fromisoformat(str(last_d)[:10])).days
            if gap > stale_days:
                return None
        except Exception:
            pass
    return float(last_v)


def _play_type(arm):
    """Revision-cycle play-type (v0 heuristic, to be refined with mcap buckets + agent judgment):
    strong upward revisions = a live catalyst (tactical); out-of-favour = mean-reversion (cyclical);
    the neutral-to-positive core = long-term compounders (structural)."""
    if arm is None:
        return "structural"
    if arm >= 75:
        return "tactical"
    if arm < 40:
        return "cyclical"
    return "structural"


# ════════════════════════════════════════════════════════════════════════════════════════════════
# FM BRAIN LIBRARY — distinct multi-force scoring philosophies (Task #85, W2-d/e)
# ════════════════════════════════════════════════════════════════════════════════════════════════
# Problem this solves: previously every virtual FM was the SAME ARM water-fill clone — only the
# mandate caps differed (`score = arm if arm is not None else 50.0`). A real FM's edge is NOVELTY:
# combining MULTIPLE forces and their interaction, not vanilla ARM/momentum. The Mesh research
# (MESH_RESEARCH_FINDINGS.md) VALIDATED that a multi-force combo beats ARM-alone walk-forward
# (OOS IC@6m: ARM 0.071 → ~0.11-0.12), and pinned the edge SOURCE: 6-month price momentum (IC@6m
# ~0.098, already > ARM by itself) + value being near-ORTHOGONAL to ARM. The defensible, non-fragile
# construction is the simple ORTHOGONALIZED ARM+momentum+value z-stack (the Σ⁻¹ optimizer and the
# regime-switch tricks add little and are unstable — decoration, dropped here).
#
# Each brain is a SCORING FUNCTION over the candidate universe (a list of {sym, isin, arm, sector,...}
# dicts). It writes u["score"] (and u["_brain_signals"] = the in-memory force z's, for play-type +
# diagnostics) IN PLACE, then the SAME waterfill() places weights under name/sector/liquidity caps —
# so ONLY the scoring changes between desks, never the construction. Signals are computed per-stock
# AS-OF the rebalance date with NO look-ahead (prices/EPS/statements ≤ asof only), reusing existing
# loaders (the price panel for momentum; data/screener bundles for value; arm.py for ARM).
#
# ── FUNDAMENTAL LAW intent (IR = IC·√BR·TC) — how each brain is DESIGNED to earn, so we can score it:
#    • IC  = where the per-bet forecast skill comes from (which force supplies the edge).
#    • √BR = how independent breadth is gained (de-correlated forces / clocks widen effective breadth).
#    • TC  = the long-only transfer leak the brain bites — turnover (momentum), capacity (deep value),
#            or spread-vs-mid-rank (quality nudge). Stated per-brain in BRAINS[...]["law"].
#
# ── LICENSING (HARD): raw per-stock LSEG StarMine ARM is used IN MEMORY ONLY for scoring. It is NEVER
#    written to a committed book/blotter/replay artifact — only DERIVED weights, the coarse play-type
#    tag, the brain-id, and AGGREGATE IC/TC ship. The existing guard (build_rules_v0 / amc_replay) is
#    untouched; these brains add no new persisted ARM column.

def _zscore(vals):
    """Cross-sectional z-score of a list (mean 0, std 1), NaN-robust: NaNs map to 0.0 (neutral).
    A list of floats in, a list of floats out (same length, same order)."""
    import math as _m
    xs = [v for v in vals if v is not None and v == v]
    if len(xs) < 2:
        return [0.0 for _ in vals]
    mu = sum(xs) / len(xs)
    var = sum((x - mu) ** 2 for x in xs) / len(xs)
    sd = _m.sqrt(var)
    if sd <= 1e-12:
        return [0.0 for _ in vals]
    return [((v - mu) / sd) if (v is not None and v == v) else 0.0 for v in vals]


def _orthogonalize(target_z, base_z):
    """Residual of `target_z` after regressing OUT `base_z` (both already z-scored, same order). Returns
    the part of target uncorrelated with base — so adding it to base supplies INDEPENDENT breadth, not a
    re-weighting of the same bet. resid = target − β·base, β = cov/var (OLS through origin on z's)."""
    n = len(target_z)
    if n < 3:
        return list(target_z)
    sxx = sum(b * b for b in base_z)
    if sxx <= 1e-12:
        return list(target_z)
    sxy = sum(t * b for t, b in zip(target_z, base_z))
    beta = sxy / sxx
    return [t - beta * b for t, b in zip(target_z, base_z)]


def momentum_6m1m(sym, asof=None):
    """Trailing 6-month price total return, SKIPPING the most recent month (classic price momentum,
    the validated edge source). = price[asof − 1m] / price[asof − 7m] − 1, read from the adjusted
    total-return panel (so it's a true TR momentum). No look-ahead: only prices ≤ asof. None if the
    7-month-ago anchor price is missing. Plain words: 'how much did this stock run over the half-year
    ending a month ago' — skipping the last month avoids the well-known 1-month reversal."""
    import pandas as pd
    df = _prices()
    if df is None or sym not in df.columns:
        return None
    s = df[sym].dropna()
    if asof:
        s = s[s.index <= str(asof)]
    if len(s) < 30:
        return None
    end = s.index[-1]
    p_recent = _asof_level(s, end - pd.DateOffset(months=1))
    p_old = _asof_level(s, end - pd.DateOffset(months=7))
    if p_recent is None or p_old is None or p_old <= 0:
        return None
    return p_recent / p_old - 1.0


def _asof_level(s, ts):
    """Last value of a price Series on/before timestamp `ts`, or None."""
    seg = s[s.index <= ts]
    return float(seg.iloc[-1]) if len(seg) else None


_SCR = {"dir": None}


def _screener_dir():
    if _SCR["dir"] is None:
        _SCR["dir"] = os.path.join(_ROOT, "data", "screener")
    return _SCR["dir"]


_VAL_CACHE = {}


def value_yields(sym, asof=None):
    """Point-in-time cheapness yields for `sym` as-of `asof`: (E/P, B/P, S/P), each higher = cheaper,
    or (None,None,None). Reads data/screener/<SYM>.json:
      • E/P = TTM-EPS-as-known-on-date / price[asof]   (EPS is an as-of-dated step series — PIT clean)
      • B/P = book-value-per-share (networth/shares, annual, lagged +3m for the results-release) / price
      • S/P = sales-per-share (annual, lagged) / price
    NO look-ahead: every fundamental carries its own report date; price is ≤ asof. Cached per (sym).
    This mirrors mesh_research._valuation_wide's conventions so the brain's value leg matches the
    validated panel. Coverage is honestly partial (not every name has a clean bundle)."""
    px = price_asof(sym, asof)
    if px is None or px <= 0:
        return (None, None, None)
    key = sym
    if key not in _VAL_CACHE:
        path = os.path.join(_screener_dir(), f"{sym}.json")
        if not os.path.exists(path):
            _VAL_CACHE[key] = None
        else:
            try:
                b = json.load(open(path, encoding="utf-8"))
            except Exception:
                b = None
            _VAL_CACHE[key] = _value_series_from_bundle(b) if b else None
    ser = _VAL_CACHE[key]
    if not ser:
        return (None, None, None)
    eps = _asof_pair(ser.get("eps"), asof)
    bvps = _asof_pair(ser.get("bvps"), asof)
    sps = _asof_pair(ser.get("sps"), asof)
    ep = (eps / px) if eps is not None else None
    bp = (bvps / px) if bvps is not None else None
    sp = (sps / px) if sps is not None else None
    return (ep, bp, sp)


def _asof_pair(pairs, asof):
    """Last value of a [(date,val),...] step series on/before `asof`, or None. (Pairs pre-sorted.)"""
    if not pairs:
        return None
    cut = str(asof) if asof else None
    last = None
    for d, v in pairs:
        if cut is None or str(d)[:10] <= cut:
            last = v
        else:
            break
    return last


def _value_series_from_bundle(b):
    """Extract per-share step series {eps, bvps, sps} from a Screener bundle — each a date-sorted
    [(asof_date, value)] list. EPS from valuation.EPS (as-of dated); BVPS/SPS from annual statements
    lagged +3m (results-release; no look-ahead). Returns {} if unextractable. Reproduces the
    mesh_research extraction so the brain's value leg == the validated panel's value leg."""
    out = {}
    val = b.get("valuation") or {}
    eps_pairs = val.get("EPS")
    if eps_pairs:
        rows = sorted((str(d), _f(v, None)) for d, v in eps_pairs if d and _f(v, None) is not None)
        if rows:
            out["eps"] = rows
    bvps, sps = _annual_bvps_sps(b)
    if bvps:
        out["bvps"] = bvps
    if sps:
        out["sps"] = sps
    return out


def _annual_bvps_sps(b):
    """Annual book-value-per-share + sales-per-share as date-sorted (asof,val) pairs from the bundle's
    profit_loss + balance_sheet statements. shares = PAT/EPS (cr); BVPS = networth/shares;
    SPS = sales/shares. as-of = fiscal-end + 3 months (results-release lag → no look-ahead)."""
    try:
        st = b.get("statements") or {}
        pl = st.get("profit_loss") if isinstance(st.get("profit_loss"), list) else None
        bs = st.get("balance_sheet") if isinstance(st.get("balance_sheet"), list) else None
        if not pl or not bs:
            return None, None
        pat = _stmt_row(pl, ["net profit"])
        eps = _stmt_row(pl, ["eps"])
        sales = _stmt_row(pl, ["sales", "revenue"])
        eqcap = _stmt_row(bs, ["equity capital"])
        reserves = _stmt_row(bs, ["reserves"])
        bvps_pairs, sps_pairs = [], []
        for c in pat:
            asof = _fiscal_asof(c)
            if asof is None:
                continue
            p, e, sa = pat.get(c), eps.get(c), sales.get(c)
            shares = (p / e) if (p is not None and e not in (None, 0)) else None
            if shares in (None, 0):
                continue
            ec, rv = eqcap.get(c), reserves.get(c)
            nw = ((ec or 0.0) + (rv or 0.0)) if (ec is not None or rv is not None) else None
            if nw is not None:
                bvps_pairs.append((asof, nw / shares))
            if sa is not None:
                sps_pairs.append((asof, sa / shares))
        bvps_pairs.sort(); sps_pairs.sort()
        return (bvps_pairs or None, sps_pairs or None)
    except Exception:
        return None, None


def _stmt_row(table, aliases):
    """{period_col: float} for the first statement row whose label matches any alias."""
    al = [a.lower() for a in aliases]
    for row in table:
        lk = "Unnamed: 0" if "Unnamed: 0" in row else next(iter(row), None)
        if lk and any(a in str(row.get(lk, "")).lower() for a in al):
            out = {}
            for c, v in row.items():
                if c == lk:
                    continue
                fv = _f(v, None)
                out[c] = fv
            return out
    return {}


def _fiscal_asof(label):
    """'Mar 2023'/'2023-03'/'Mar-23' → 'YYYY-MM-DD' string = fiscal-end + 3 months (results lag), or
    None. Conservative (a year's figure is only 'known' a quarter after year-end → no look-ahead)."""
    import pandas as pd
    ts = pd.to_datetime(str(label).strip(), errors="coerce")
    if pd.isna(ts):
        for fmt in ("%b %Y", "%b-%y", "%b-%Y", "%Y-%m", "%Y"):
            try:
                ts = pd.Timestamp(pd.to_datetime(str(label).strip(), format=fmt))
                break
            except Exception:
                continue
    if pd.isna(ts):
        return None
    return ((ts + pd.offsets.MonthEnd(0) + pd.DateOffset(months=3)) + pd.offsets.MonthEnd(0)).strftime("%Y-%m-%d")


# ── the brains ────────────────────────────────────────────────────────────────────────────────────
# Each takes (uni, asof) where uni = list of candidate dicts (mutated in place: u["score"] set, and
# u["arm"] expected present). Returns a small diag dict (force coverage, the brain id). After a brain
# runs, the SAME waterfill places weights — only the score changed.

def _attach_forces(uni, asof):
    """Compute the cross-sectional force z-scores ONCE for the candidate set as-of `asof`, and attach
    them to each u: u["z_arm"], u["z_mom"], u["z_val"] (each 0.0 where the raw force is missing — a
    neutral nudge, never a penalty). Returns coverage counts. Shared by every brain so the signals are
    computed identically across desks (the only difference is how a brain WEIGHTS these three z's)."""
    arm_raw = [u.get("arm") for u in uni]
    mom_raw = [momentum_6m1m(u["sym"], asof) for u in uni]
    val_raw = []
    for u in uni:
        ep, bp, sp = value_yields(u["sym"], asof)
        legs = [_zsafe(x) for x in (ep, bp, sp)]
        # combine the three cheapness yields cross-sectionally below; store the raw triple for now
        u["_val_legs"] = (ep, bp, sp)
    z_arm = _zscore(arm_raw)
    z_mom = _zscore(mom_raw)
    # value_z = mean of the cross-sectional z of E/P, B/P, S/P (matches mesh_research._combined_value_z)
    z_ep = _zscore([u["_val_legs"][0] for u in uni])
    z_bp = _zscore([u["_val_legs"][1] for u in uni])
    z_sp = _zscore([u["_val_legs"][2] for u in uni])
    z_val = []
    for i, u in enumerate(uni):
        legs = [z for z, raw in zip((z_ep[i], z_bp[i], z_sp[i]), u["_val_legs"]) if raw is not None and raw == raw]
        z_val.append(sum(legs) / len(legs) if legs else 0.0)
    for i, u in enumerate(uni):
        u["z_arm"], u["z_mom"], u["z_val"] = z_arm[i], z_mom[i], z_val[i]
        u.pop("_val_legs", None)
    return {
        "n": len(uni),
        "cov_arm": sum(1 for r in arm_raw if r is not None),
        "cov_mom": sum(1 for r in mom_raw if r is not None and r == r),
        "cov_val": sum(1 for u in uni if u["z_val"] != 0.0),
    }


def _zsafe(x):
    return x if (x is not None and x == x) else None


def brain_core_multifactor(uni, asof):
    """LENS 1 — CORE-MULTIFACTOR (the validated default workhorse).
    score = z(ARM) + z(resid_mom⊥ARM) + 0.5·z(resid_val⊥ARM). Momentum and value are ORTHOGONALIZED
    against ARM so each adds INDEPENDENT breadth (not a re-weighting of revisions). This is the one
    construction the research stress-tested to destruction (OOS IC@6m ~0.11, 100% all-starts beat-rate).
    LAW: IC from two replicated, imperfectly-correlated factors (analyst-revision + price momentum);
    BR widened by adding near-orthogonal value; TC leak = momentum TURNOVER (the binding constraint —
    its deployable IR must be turnover-validated)."""
    cov = _attach_forces(uni, asof)
    z_arm = [u["z_arm"] for u in uni]
    r_mom = _orthogonalize([u["z_mom"] for u in uni], z_arm)
    r_val = _orthogonalize([u["z_val"] for u in uni], z_arm)
    for i, u in enumerate(uni):
        u["score_raw"] = z_arm[i] + r_mom[i] + 0.5 * r_val[i]
    _rescale_positive(uni)
    cov["brain"] = "core_multifactor"
    return cov


def brain_momentum_led(uni, asof):
    """LENS 2 — MOMENTUM-LED (price momentum primary, ARM/value as confirmation/filter).
    score = z(mom) + 0.5·z(resid_arm⊥mom) + 0.25·z(resid_val⊥mom). Momentum leads (it's the single
    strongest standalone force, IC@6m ~0.098); ARM confirms the trend has analyst support; value is a
    light contrarian brake on chasing the most expensive winners.
    LAW: IC primarily from price momentum; BR from the ARM-confirmation overlay (de-correlated wins);
    TC leak = HIGH TURNOVER (fastest clock of the four brains) — the hardest transfer, best where
    liquidity is deep (large/mid)."""
    cov = _attach_forces(uni, asof)
    z_mom = [u["z_mom"] for u in uni]
    r_arm = _orthogonalize([u["z_arm"] for u in uni], z_mom)
    r_val = _orthogonalize([u["z_val"] for u in uni], z_mom)
    for i, u in enumerate(uni):
        u["score_raw"] = z_mom[i] + 0.5 * r_arm[i] + 0.25 * r_val[i]
    _rescale_positive(uni)
    cov["brain"] = "momentum_led"
    return cov


def brain_value_revision(uni, asof):
    """LENS 3 — VALUE-REVISION CONTRARIAN (cheap AND being upgraded).
    score = z(val) + z(ARM) (additive rank-sum — the validated form; the multiplicative 'corner' is
    DEAD). Long names that are both cheap and seeing rising analyst estimates — a revision turning up
    ON a cheap name is the catalyst that prevents a value trap.
    LAW: IC from value DIVERSIFYING ARM (margin +0.019, but explicitly regime-conditional — value
    HURTS in growth-led 2017-20-type regimes; that's its identity, not a bug); BR from value⊥revision
    near-orthogonality (two real independent legs); TC leak = CAPACITY (deep-value names are often less
    liquid → a long-only book can't always size the cheapest names). Lowest turnover, best price-TC,
    but a capacity/liquidity leak instead."""
    cov = _attach_forces(uni, asof)
    for u in uni:
        u["score_raw"] = u["z_val"] + u["z_arm"]
    _rescale_positive(uni)
    cov["brain"] = "value_revision"
    return cov


def brain_regime_switch(uni, asof, regime=None):
    """LENS 4 — REGIME-SWITCH (tilt the blend by a simple market read).
    A single breadth/trend regime read decides the mix between the other brains:
      • RISK-ON  (broad market uptrend, breadth strong) → momentum-led tilt (ride the trend).
      • RISK-OFF (market below trend, breadth weak)      → value-revision tilt (defensive/cheap).
      • NEUTRAL                                          → the core multi-factor blend.
    The regime is read from the candidate set's OWN cross-section as-of date (no external macro call,
    no look-ahead): fraction of names with positive 6m momentum (breadth-of-momentum) — a cheap, robust
    market-state proxy. > 0.55 = risk-on, < 0.45 = risk-off, else neutral.
    LAW: IC engine ROTATES with the state (momentum's IC dominates in trends, value's in
    mean-reversion) — this is a TC/BR management lens more than a new IC source: it buys the right
    transfer leak for the regime (turnover when trends pay, capacity-patience when they don't). The
    research flagged a HARD regime-switch as fragile, so this is a SOFT 3-state tilt over the SAME
    validated forces, not a hand-tuned weight vector."""
    cov = _attach_forces(uni, asof)
    if regime is None:
        regime = _read_regime(uni)
    z_arm = [u["z_arm"] for u in uni]
    z_mom = [u["z_mom"] for u in uni]
    z_val = [u["z_val"] for u in uni]
    if regime == "risk_on":           # momentum-led
        r_arm = _orthogonalize(z_arm, z_mom)
        r_val = _orthogonalize(z_val, z_mom)
        sc = [z_mom[i] + 0.5 * r_arm[i] + 0.25 * r_val[i] for i in range(len(uni))]
    elif regime == "risk_off":        # value-revision
        sc = [z_val[i] + z_arm[i] for i in range(len(uni))]
    else:                             # neutral → core multifactor
        r_mom = _orthogonalize(z_mom, z_arm)
        r_val = _orthogonalize(z_val, z_arm)
        sc = [z_arm[i] + r_mom[i] + 0.5 * r_val[i] for i in range(len(uni))]
    for i, u in enumerate(uni):
        u["score_raw"] = sc[i]
    _rescale_positive(uni)
    cov["brain"] = "regime_switch"
    cov["regime"] = regime
    return cov


def _read_regime(uni):
    """Cheap, no-look-ahead market-state read = breadth of 6-month momentum across the candidate set
    (fraction with z_mom > 0). > 0.55 risk_on, < 0.45 risk_off, else neutral. (z_mom is mean-zero by
    construction, so 'fraction > 0' is a clean breadth proxy of which way the cross-section is tilted.)"""
    z = [u.get("z_mom", 0.0) for u in uni]
    if not z:
        return "neutral"
    frac_up = sum(1 for v in z if v > 0) / len(z)
    return "risk_on" if frac_up > 0.55 else ("risk_off" if frac_up < 0.45 else "neutral")


def _rescale_positive(uni):
    """The waterfill places weight ∝ score, so scores must be POSITIVE. Map each brain's signed
    score_raw onto a strictly-positive scale by SHIFTING above the minimum (preserves the full ranking
    AND the relative spacing). score = score_raw − min + 1 (so the worst name still gets a small
    positive base weight, the best gets the largest)."""
    raws = [u.get("score_raw", 0.0) for u in uni]
    if not raws:
        return
    lo = min(raws)
    for u in uni:
        u["score"] = (u.get("score_raw", 0.0) - lo) + 1.0


# ── brain registry + mandate→brain assignment ──────────────────────────────────────────────────────
# Each entry: fn = the scoring function; play = default play-type horizon the brain trades on;
# thesis = one-line plain-English brain; law = the IC·√BR·TC design intent (how we'll SCORE the desk).
BRAINS = {
    "core_multifactor": {
        "fn": brain_core_multifactor, "play": "structural",
        "thesis": "Analyst upgrades + price momentum + a cheapness lean, combined as orthogonalized "
                  "residuals — the validated default that captures the one real multi-force edge.",
        "law": "IC: two replicated imperfectly-correlated factors (ARM + 6m momentum). "
               "BR: widened by near-orthogonal value. TC leak: momentum turnover (the binding cost).",
    },
    "momentum_led": {
        "fn": brain_momentum_led, "play": "tactical",
        "thesis": "Ride the strongest trends (6m price momentum) but only where analysts confirm and "
                  "the name isn't egregiously expensive — a fast, trend-following desk.",
        "law": "IC: price momentum (strongest standalone force). BR: ARM-confirmation overlay. "
               "TC leak: HIGH turnover (fastest clock) — needs deep liquidity (large/mid).",
    },
    "value_revision": {
        "fn": brain_value_revision, "play": "cyclical",
        "thesis": "Buy cheap names whose analyst estimates have started turning UP — cheapness plus a "
                  "revision catalyst, the contrarian's edge against the value trap.",
        "law": "IC: value diversifying ARM (regime-conditional — loses in growth-led regimes by "
               "design). BR: value⊥revision orthogonality. TC leak: capacity (deep-value illiquidity).",
    },
    "regime_switch": {
        "fn": brain_regime_switch, "play": "structural",
        "thesis": "Tilt between momentum (risk-on) and value-revision (risk-off) by a simple "
                  "breadth-of-momentum market read; the core blend when the state is unclear.",
        "law": "IC: rotates with the regime (momentum in trends, value in reversals). "
               "BR/TC: a transfer-management lens — buys the right leak for the state, soft 3-state.",
    },
}
_DEFAULT_BRAIN = "core_multifactor"


def brain_for_mandate(category, mandate=None):
    """Assign a DISTINCT FM brain to a scheme by its SEBI mandate, so different desks run different
    philosophies (not one ARM clone):
      • value / contra / dividend tilt  → value_revision  (cheap + upgrading — the stated style)
      • hybrid / balanced-adv / DAA / multi-asset (an `equity_max` band = actively FLEXES equity)
                                        → regime_switch   (the mandate IS a market-state allocator)
      • small / mid / sectoral-thematic → momentum_led    (trend persistence strongest + tradable)
      • everything else (large / flexi / multi / focused / ELSS) → core_multifactor (the default)
    Returns a brain-id present in BRAINS."""
    m = mandate or {}
    cat = str(category or "").lower()
    if m.get("tilt") in ("value", "yield") or "value" in cat or "contra" in cat:
        return "value_revision"
    # any mandate with an equity band it must flex (hybrid / balanced-advantage / DAA / multi-asset /
    # equity-savings) is, by design, a market-state allocator → the regime-switch desk.
    if m.get("daa") or m.get("equity_max") or "balanced advantage" in cat or "dynamic asset" in cat:
        return "regime_switch"
    buckets = set(m.get("buckets", []))
    if m.get("thematic") or "sectoral" in cat or "thematic" in cat:
        return "momentum_led"
    if buckets == {"small"} or buckets == {"mid"}:
        return "momentum_led"
    return _DEFAULT_BRAIN


def score_universe(uni, asof, brain_id=None):
    """Run the chosen FM brain over the candidate universe `uni` (list of dicts), writing u["score"]
    in place. Returns (brain_id, diag). Falls back to the default brain on an unknown id. THE single
    entry point both build_rules_v0 (live book) and amc_replay.construct_targets (walk-forward) call —
    so the scoring philosophy is shared and only the brain-id differs between desks."""
    bid = brain_id if brain_id in BRAINS else _DEFAULT_BRAIN
    diag = BRAINS[bid]["fn"](uni, asof)
    diag["brain"] = bid
    return bid, diag


def waterfill(sel, equity_target, max_sector, sector_free="Unclassified", iters=50):
    """Place `equity_target` (fraction of NAV) across the names in `sel` ∝ each name's `score`,
    honoring per-name caps (`u['cap']`) AND a per-sector ceiling `max_sector` (the `sector_free`
    pseudo-sector is exempt — unknown-sector names aren't really one sector). Leftover weight from
    names that hit a cap redistributes to those with headroom over `iters` passes; whatever still
    can't fit (a genuine capacity limit) is left undeployed. Mutates `u['w']` in place; returns the
    total fraction actually deployed. THE single water-fill used by both build_rules_v0 (today's
    book) and amc_replay (the historical walk-forward) so they stay numerically identical."""
    for u in sel:
        u["w"] = 0.0
    remaining = equity_target
    for _ in range(iters):
        if remaining <= 1e-4:
            break
        secw = {}
        for u in sel:
            secw[u["sector"]] = secw.get(u["sector"], 0.0) + u["w"]
        active = [u for u in sel if u["w"] < u["cap"] - 1e-9 and
                  (u["sector"] == sector_free or secw.get(u["sector"], 0.0) < max_sector - 1e-9)]
        if not active:
            break
        tot_score = sum(x["score"] for x in active) or 1.0
        placed = 0.0
        for u in active:
            sec_room = 1e9 if u["sector"] == sector_free else max_sector - secw.get(u["sector"], 0.0)
            add = min(remaining * u["score"] / tot_score, u["cap"] - u["w"], sec_room)
            if add > 0:
                u["w"] += add
                secw[u["sector"]] = secw.get(u["sector"], 0.0) + add
                placed += add
        remaining -= placed
        if placed <= 1e-7:
            break
    return equity_target - remaining


def build_rules_v0(reg_entry, asof, equity_target=0.95, brain_id=None, log=print):
    """Deploy a fresh 100%-cash book into a mandate+liquidity-constrained, MULTI-FORCE-scored
    portfolio as-of `asof`. Returns (book, trades, diag). No look-ahead: only prices/ARM/EPS ≤ asof.

    Method (reproducible): universe = the scheme's disclosed equity holdings that we can price on
    `asof`; the chosen FM BRAIN (brain_for_mandate, or `brain_id` override) scores the set by a
    DISTINCT multi-force philosophy (ARM + momentum + value, combined per the brain — NOT a single
    ARM clone); pick the top min(n_hi, |U|) by that score; water-fill weights ∝ score under the
    mandate single-name + sector caps and the liquidity cap; whatever isn't deployed stays cash.

    The book/trades are tagged with the brain-id + its play-type so the desk's philosophy is legible.
    LICENSING: raw per-stock ARM is used only IN MEMORY to score — only derived weights / play-type /
    brain-id / aggregate coverage are persisted (the guard below is unchanged)."""
    m = reg_entry["mandate"]
    aum = _f(reg_entry["aum_cr"])
    asof = str(asof)
    bid = brain_id or brain_for_mandate(reg_entry.get("category"), m)
    # a hybrid / balanced-advantage mandate caps equity (equity_max) — the rest is the debt+cash
    # sleeve we don't model as stocks, so it shows as cash. Target the middle of the equity band.
    emax = m.get("equity_max")
    if emax:
        equity_target = min(equity_target, round((m["equity_min"] + emax) / 2.0, 4))

    uni = []
    n_priced = n_arm = 0
    for h in reg_entry.get("real_holdings", []):
        sym = h.get("symbol")
        px = price_asof(sym, asof) if sym else None
        if not sym or px is None or px <= 0:
            continue
        n_priced += 1
        arm = current_arm(h.get("isin"), asof)
        if arm is not None:
            n_arm += 1
        sec = (h.get("industry") or h.get("sector") or "").strip()
        if sec.upper() in ("", "N.A.", "NA", "-", "NONE", "UNCLASSIFIED"):
            sec = "Unclassified"
        uni.append({"sym": sym, "isin": h.get("isin"), "name": h.get("name") or sym,
                    "sector": sec, "px": px, "arm": arm, "real_pct": _f(h.get("pct"))})
    if not uni:
        raise SystemExit("no priceable holdings — cannot construct a book")

    # ── SCORE with the desk's FM brain (multi-force: ARM + momentum + value, combined per the brain)
    bid, brain_diag = score_universe(uni, asof, bid)

    uni.sort(key=lambda u: (-u["score"], -u["real_pct"]))
    sel = uni[:min(m["n_hi"], len(uni))]
    SECTOR_FREE = "Unclassified"          # unknown-sector names aren't really one sector → no sector cap

    # per-name cap = min(mandate single-name cap, liquidity cap as a fraction of AUM)
    for u in sel:
        cap = m["max_pos"]
        lc = liquidity_cap_cr(u["sym"], aum, asof)
        if lc is not None and aum > 0:
            cap = min(cap, lc / aum)
        u["cap"] = max(0.0, cap)
        u["w"] = 0.0

    # iterative water-fill: place equity_target across names ∝ score, honoring per-name caps AND
    # per-sector caps; leftover from capped names redistributes to those with headroom. Whatever
    # still can't fit (a genuine capacity limit — e.g. a huge book in illiquid small-caps) stays cash.
    waterfill(sel, equity_target, m["max_sector"], SECTOR_FREE)

    book = new_book(reg_entry)
    book["inception"] = asof
    book["asof"] = asof
    book["brain"] = bid                              # the desk's FM philosophy (derived tag, not ARM)
    book["brain_thesis"] = BRAINS[bid]["thesis"]
    # NOTE on what gets PERSISTED: the raw per-stock ARM value is LICENSED LSEG IP and must NEVER
    # be written to a committed file — so the book/blotter store only OUR derived decision (weight,
    # within-book selection rank, the coarse play-type horizon tag, the brain-id) and a qualitative
    # rationale. The exact ARM score (and the per-stock force z's) stay in-memory (diag) and are
    # reproducible locally from arm_repo + the price/screener panels.
    trades, deployed = [], 0.0
    play_default = BRAINS[bid]["play"]
    for rank, u in enumerate([x for x in sel if x["w"] > 0], 1):
        qty = round(u["w"] * aum * 1e7 / u["px"])
        if qty <= 0:
            continue
        play = _play_type(u["arm"])
        book["positions"][u["sym"]] = {
            "isin": u["isin"], "name": u["name"], "sector": u["sector"],
            "qty": qty, "avg_cost": round(u["px"], 4), "play_type": play,
            "entry_date": asof, "thesis_ref": None, "sel_rank": rank, "brain": bid,
        }
        val = qty * u["px"] / 1e7
        deployed += val
        trades.append({"date": asof, "sym": u["sym"], "isin": u["isin"], "name": u["name"],
                       "side": "BUY", "qty": qty, "price": round(u["px"], 4), "value_cr": round(val, 4),
                       "play_type": play, "brain": bid,
                       "rationale": f"quant rank #{rank} ({bid}); deploy {round(u['w']*100,2)}% as "
                       f"{play} under {m['max_pos']*100:.0f}% name / {m['max_sector']*100:.0f}% sector "
                       f"/ liquidity caps"})
    book["cash_cr"] = round(aum - deployed, 4)
    diag = {"universe": len(uni), "n_priced": n_priced, "n_arm_scored": n_arm,
            "brain": bid, "brain_coverage": {k: brain_diag.get(k) for k in ("cov_arm", "cov_mom", "cov_val")},
            "regime": brain_diag.get("regime"),
            "selected": len(book["positions"]), "deployed_cr": round(deployed, 1),
            "deployed_pct": round(100 * deployed / aum, 2) if aum else None}
    log(f"  built book [{bid}]: {diag['selected']} names, deployed {diag['deployed_pct']}% "
        f"(₹{diag['deployed_cr']:,} cr), cash {round(100-diag['deployed_pct'],2)}%; forces "
        f"ARM {brain_diag.get('cov_arm')}/mom {brain_diag.get('cov_mom')}/val {brain_diag.get('cov_val')}"
        f" of {n_priced} priced names")
    return book, trades, diag


# ───────────────────────────────────────────────────────── .xlsx render (ABSL CITI layout)
def to_xlsx(sheet, path):
    """Render a fact_sheet() dict to an .xlsx mirroring ABSL's CITI_BIRLA_DAILY_EQUITY_FACT_SHEET:
    header block, sector-grouped holdings with subtotals, and a footer (equity/cash/NAV/day-return)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Equity Fact Sheet"
    bold = Font(bold=True)
    white = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    sub_fill = PatternFill("solid", fgColor="D9E1F2")
    foot_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    h = sheet["header"]
    ws["A1"] = "VIRTUAL DAILY EQUITY FACT SHEET"; ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"{h.get('amc')}  —  {h.get('scheme')}"; ws["A2"].font = bold
    ws["A3"] = f"As of {h.get('asof')}    AUM (mark-to-market): ₹{h.get('aum_cr'):,.2f} cr"
    cols = ["SR", "ISIN", "NAME", "SECTOR", "PLAY", "QTY", "AVG COST ₹", "BOOK COST ₹cr",
            "MKT PRICE ₹", "MKT VALUE ₹cr", "PREV PRICE ₹", "PREV VALUE ₹cr",
            "% CHG", "WTD CONTRIB", "% ASSETS"]
    r0 = 5
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=r0, column=j, value=c)
        cell.font = white; cell.fill = hdr_fill; cell.border = box; cell.alignment = right
    ws.cell(row=r0, column=1).alignment = Alignment(horizontal="left")

    # group rows by sector (rows already sorted by sector then value in fact_sheet)
    sub = {s["sector"]: s for s in sheet["sectors"]}
    r = r0 + 1
    last_sec = None
    num = {6: "#,##0", 7: "#,##0.00", 8: "#,##0.000", 9: "#,##0.00", 10: "#,##0.000",
           11: "#,##0.00", 12: "#,##0.000", 13: "0.00", 14: "0.000", 15: "0.00"}
    for row in sheet["rows"]:
        if row["sector"] != last_sec:
            if last_sec is not None:
                _xlsx_subtotal(ws, r, last_sec, sub.get(last_sec), sub_fill, bold, box); r += 1
            last_sec = row["sector"]
        vals = [row["sr"], row["isin"], row["name"], row["sector"], row.get("play_type"),
                row["qty"], row["avg_cost"], row["book_cost"], row["mkt_price"], row["mkt_value"],
                row["prev_price"], row["prev_value"], row["pct_change"], row["wtd_contribution"],
                row["pct_assets"]]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = box
            if j in num and v is not None:
                cell.number_format = num[j]
        r += 1
    if last_sec is not None:
        _xlsx_subtotal(ws, r, last_sec, sub.get(last_sec), sub_fill, bold, box); r += 1

    f = sheet["footer"]
    r += 1
    foot = [("EQUITY", f["equity_cr"], "₹cr"), ("CASH", f["cash_cr"], f"₹cr ({f.get('cash_pct')}%)"),
            ("TOTAL (NAV base)", f["total_cr"], "₹cr"), ("HOLDINGS", f["n_holdings"], "names"),
            ("DAY RETURN", f.get("day_return_pct"), "%")]
    for label, v, unit in foot:
        ws.cell(row=r, column=2, value=label).font = bold
        c = ws.cell(row=r, column=3, value=v); c.fill = foot_fill; c.font = bold
        ws.cell(row=r, column=4, value=unit)
        r += 1

    widths = [5, 15, 30, 22, 11, 12, 12, 13, 12, 13, 12, 13, 8, 11, 9]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=r0, column=j).column_letter].width = w
    ws.freeze_panes = "A6"
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    return path


def _xlsx_subtotal(ws, r, sector, s, fill, bold, box):
    ws.cell(row=r, column=3, value=f"{sector} — subtotal").font = bold
    if s:
        ws.cell(row=r, column=10, value=round(s["mkt_value"], 3)).number_format = "#,##0.000"
        ws.cell(row=r, column=15, value=round(s["pct_assets"], 2)).number_format = "0.00"
    for j in range(1, 16):
        cell = ws.cell(row=r, column=j); cell.fill = fill; cell.border = box


if __name__ == "__main__":
    import sys
    args = sys.argv[1:]
    PILOT_AMCS = ["ICICI Prudential", "SBI Mutual", "Aditya Birla Sun Life", "Quant Mutual"]
    reg = registry(amcs=PILOT_AMCS, min_aum_cr=1.0)
    print(f"registry: {len(reg)} AMCs, {sum(len(v) for v in reg.values())} schemes")
    for amc, schemes in reg.items():
        big = sorted(schemes.values(), key=lambda s: -_f(s["aum_cr"]))[:4]
        print(f"\n{amc}: {len(schemes)} schemes")
        for s in big:
            print(f"   {(s['scheme'] or '?')[:46]:<46} {s['category'][:26]:<26} "
                  f"AUM ₹{s['aum_cr']:>10,.0f} cr  bench={s['benchmark']}  realIR={s['real_ir']}")
