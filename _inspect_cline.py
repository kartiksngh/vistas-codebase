"""Understand the Cline (Capitaline) monthly MF portfolio file — structure first, small & thorough."""
import sys, openpyxl
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\Administrator\Documents\Projects\MoneyBall\Cline Data on Portfolio, NAV\Cline portfolios July'25 to May'26\MF Data - May26.xlsx"

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("SHEETS:", wb.sheetnames)
print()

for sn in wb.sheetnames:
    ws = wb[sn]
    # read_only: iterate to count rows, cap to avoid huge scans but report if capped
    nrows = 0
    first_rows = []
    maxcols = 0
    for r in ws.iter_rows(values_only=True):
        if nrows < 30:
            first_rows.append(r)
            maxcols = max(maxcols, len(r))
        nrows += 1
        if nrows >= 200000:
            break
    print(f"==================== SHEET: {sn!r}  rows={nrows}{'+' if nrows>=200000 else ''}  cols~={maxcols} ====================")
    for i, r in enumerate(first_rows):
        cells = []
        for c in r[:maxcols]:
            s = "" if c is None else str(c)
            if len(s) > 28:
                s = s[:25] + "..."
            cells.append(s)
        print(f"  r{i:02d}: " + " | ".join(cells))
    print()
wb.close()
