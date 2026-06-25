"""Inspect the 3 BBG files (ISIN Map, Market Cap, Prices): schema, wide-vs-long, span, id columns."""
import openpyxl
BASE = r"C:\Users\Administrator\Documents\Projects\MoneyBall\Cline Data on Portfolio, NAV\Update December 2025\Portfolio Update\BBG Data"
FILES = {
 "ISIN_Map":   BASE + r"\ISIN Map Cline Portfolio Data Jan'10 to Oct'25 - updated Dec'25.xlsx",
 "MarketCap":  BASE + r"\Market Cap Cline Portfolio Data Jan'10 to Oct'25 - updated Dec'25.xlsx",
 "Prices":     BASE + r"\Prices Cline Portfolio Data Jan'10 to Oct'25 - updated Dec'25.xlsx",
}
for tag, path in FILES.items():
    print(f"\n{'='*72}\n{tag}\n{'='*72}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print("  open failed:", e); continue
    print("SHEETS:", wb.sheetnames)
    for sn in wb.sheetnames[:3]:
        ws = wb[sn]
        # sample first rows quickly
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append(r)
            if i >= 11: break
        ncol = max((len(r) for r in rows), default=0)
        print(f"\n  --- sheet {sn!r}  sampled_cols~={ncol} ---")
        for i, r in enumerate(rows):
            cells = []
            for c in r[:14]:
                s = "" if c is None else str(c)
                cells.append(s[:18])
            tail = ""
            if len(r) > 14:
                # show a few of the LAST columns (date span if wide)
                last = [("" if c is None else str(c))[:12] for c in r[-3:]]
                tail = "  ...  " + " | ".join(last)
            print(f"   r{i:02d}: " + " | ".join(cells) + tail)
    wb.close()
